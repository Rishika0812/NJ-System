"""
S³ Export — Momentum INTERACTIVE Workbook
=========================================
A single control cell drives the whole workbook: type a cycle number on the
Dashboard and every "View" sheet (Trades, Eligible Ranking, Candidates, Leg
Rankings, NIFTY Path, Monthwise) auto-filters to that cycle via live formulas.

Requires Excel 2021/365 or LibreOffice Calc (uses dynamic-array FILTER).
Aggregate cards use SUMIFS/COUNTIFS/AVERAGEIFS and work everywhere.
"""
from __future__ import annotations

import io
from copy import copy
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from core.investment_analysis import compute_investment_analysis
    from export.momentum_exporter import generate_momentum_excel
except Exception:  # pragma: no cover
    from investment_analysis import compute_investment_analysis  # type: ignore
    from momentum_exporter import generate_momentum_excel  # type: ignore

_FONT = "Arial"
_HEAD_BG = "2E4057"; _HEAD_FG = "FFFFFF"
_TITLE_BG = "6C63FF"; _SUB_BG = "44476A"
_GREEN_BG = "C6EFCE"; _GREEN_FG = "1A6B3C"

def _make_run_name(cfg: dict) -> str:
    """Compact settings-based run name for titles and the workbook."""
    m   = cfg.get("metric", "roc")
    vt  = cfg.get("vol_type", "standard")
    vd  = cfg.get("vol_dir", cfg.get("vol_direction", "low"))
    side = cfg.get("selection_side", "top")
    top_n = cfg.get("top_n", cfg.get("vol_filter_n", 50))
    top_k = cfg.get("top_k", 10)
    use_ff = cfg.get("use_fixed_fall", False)
    fe_d  = cfg.get("fall_entry_days", 100)
    fx_d  = cfg.get("fall_exit_days", 100)
    np_   = cfg.get("nifty_pct", 5)
    fp_   = cfg.get("fall_pct", 10)
    hd_   = cfg.get("max_hold", 2)
    exact = cfg.get("exact_trigger_mode", False)
    _ds = "DS" if vt == "downside" else "Std"
    _lo = "Lo" if vd == "low" else "Hi"
    m_tok = {
        "roc":         "ROC",
        "vol":         f"Vol-{_ds}-{_lo}",
        "volxroc":     f"RxV-{_ds}",
        "roc_over_vol":f"R/V-{_ds}-{_lo}",
        "both":        f"Both-{_ds}-{_lo}",
        "beta":        "Beta",
        "beta_over_vol": f"BetaOVol-{_ds}",
        "beta_x_vol":  f"BetaXVol-{_ds}",
        "sd_over_dv":  "SdOvDV",
        "gate":        "Gate-ARQM",
        "off":         f"Off-{_ds}-{_lo}",
    }.get(m, m.upper())
    win_tok  = f"FF{fe_d}+{fx_d}d" if use_ff else "AutoLegs"
    side_tok = "Bot" if side == "bottom" else "Top"
    exact_tok = " ExactTrig" if exact else ""
    return (f"{m_tok} {side_tok}{top_n} K{top_k} {win_tok} "
            f"N{np_:g}% Fall{fp_:g}% Hold{hd_:g}yr{exact_tok}")

_RED_BG = "FFC7CE"; _RED_FG = "9C0006"
_GOLD = "FFF2CC"; _CTRL = "FFE699"; _ALT = "F2F4FF"; _BORDER = "B8C2D9"


def _thin():
    s = Side(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_date(d):
    try:
        if d is None or pd.isna(d):
            return ""
    except Exception:
        pass
    try:
        return pd.Timestamp(d).strftime("%d-%b-%Y")
    except Exception:
        return str(d) if d is not None else ""


def _co(v):
    if isinstance(v, pd.Timestamp):
        return _fmt_date(v)
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
        return ""
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return round(float(v), 4)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    return v


def _title(ws, text, ncols, bg=_TITLE_BG):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, ncols))
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=_FONT, bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24


def _band(ws, row, text, ncols, bg=_SUB_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(1, ncols))
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=_FONT, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 19


def _df_to_table(wb, sheet_name, df: pd.DataFrame, table_name: str, hidden=True,
                 key_col="Cycle"):
    """Write a DataFrame as a real Excel Table (header row 1) on its own sheet."""
    ws = wb.create_sheet(sheet_name)
    if df is None or df.empty:
        df = pd.DataFrame({key_col: []})
    else:
        # Drop columns that are completely blank/null for this setting
        valid_cols = []
        for c in df.columns:
            s = df[c].dropna()
            if not s.empty:
                is_blank = s.map(lambda v: str(v).strip() in ("", "—", "nan", "None", "<NA>")).all()
                if not is_blank:
                    valid_cols.append(c)
            elif c in (key_col, "Ticker", "Cycle"):
                valid_cols.append(c)
        if valid_cols:
            df = df[valid_cols]
    if key_col in df.columns and len(df):
        df = df.sort_values(key_col, kind="stable").reset_index(drop=True)
    cols = list(df.columns)
    # Header row
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=str(c))
        cell.font = Font(name=_FONT, bold=True, size=9, color=_HEAD_FG)
        cell.fill = PatternFill("solid", fgColor=_HEAD_BG)
    # Data rows — use ws.append() instead of cell-by-cell (10x faster)
    nrow = max(len(df), 1)
    for i in range(len(df)):
        ws.append([_co(df.iloc[i][c]) for c in cols])
    if len(df) == 0:
        ws.append([None] * len(cols))  # ensure table has ≥1 data row
    ref = f"A1:{get_column_letter(len(cols))}{1 + nrow}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True,
                                        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    if hidden:
        ws.sheet_state = "hidden"
    return ws


def _filter_view(ws, start_row, display_headers, table_name, col_names, kmax,
                 ctrl="Dashboard!$C$2", green_col=None, traded_col=None, key_col="Cycle"):
    """Header row + non-array MATCH/COUNTIF/INDEX formulas that show ONLY the rows
    of `table_name` whose key (`key_col`) == the control cell (universally compatible).

    `col_names` are the table's column tokens (in display order); `display_headers`
    are the friendly labels. `kmax` = max rows any single key value can have.
    """
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    border = _thin()
    ncol = len(display_headers)
    # helper cells (off to the right): start row + count for the selected key
    hcol = ncol + 2
    hs = get_column_letter(hcol)
    ws.cell(row=start_row, column=hcol, value=f'=IFERROR(MATCH({ctrl},{table_name}[{key_col}],0),0)')
    ws.cell(row=start_row + 1, column=hcol, value=f'=COUNTIF({table_name}[{key_col}],{ctrl})')
    ws.cell(row=start_row, column=hcol + 1, value="◀ start / count (helper)")\
        .font = Font(name=_FONT, italic=True, size=8, color="999999")
    sref = f"${hs}${start_row}"; cref = f"${hs}${start_row + 1}"

    for j, h in enumerate(display_headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(name=_FONT, bold=True, size=10, color=_HEAD_FG)
        c.fill = PatternFill("solid", fgColor=_HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(j)].width = 15

    data_top = start_row + 1
    for k in range(1, kmax + 1):
        r = start_row + k
        for j, col in enumerate(col_names, start=1):
            f = (f'=IF((ROW()-{start_row})<={cref},'
                 f'IFERROR(INDEX({table_name}[{col}],{sref}+ROW()-{start_row}-1),""),"")')
            cell = ws.cell(row=r, column=j, value=f)
            cell.font = Font(name=_FONT, size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    data_bot = start_row + kmax
    rng = f"A{data_top}:{get_column_letter(ncol)}{data_bot}"

    # dynamic highlight: green/red on a return column
    if green_col and green_col in display_headers:
        gi = display_headers.index(green_col) + 1
        gl = get_column_letter(gi)
        gcell_top = f"{gl}{data_top}"
        ws.conditional_formatting.add(
            f"{gl}{data_top}:{gl}{data_bot}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=_GREEN_BG),
                       font=Font(color=_GREEN_FG)))
        ws.conditional_formatting.add(
            f"{gl}{data_top}:{gl}{data_bot}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=_RED_BG),
                       font=Font(color=_RED_FG)))
    # dynamic highlight: whole row green when Traded = "Y"
    if traded_col and traded_col in display_headers:
        ti = display_headers.index(traded_col) + 1
        tl = get_column_letter(ti)
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${tl}{data_top}="Y"'],
                        fill=PatternFill("solid", fgColor=_GOLD)))


def _kv(ws, row, pairs, label_w=26, val_w=22, val_col=2):
    for k, v in pairs:
        a = ws.cell(row=row, column=1, value=k)
        a.font = Font(name=_FONT, bold=True, size=10)
        a.alignment = Alignment(horizontal="left", indent=1)
        b = ws.cell(row=row, column=val_col, value=v)
        b.font = Font(name=_FONT, size=10)
        b.alignment = Alignment(horizontal="left", indent=1)
        row += 1
    ws.column_dimensions["A"].width = label_w
    ws.column_dimensions[get_column_letter(val_col)].width = val_w
    return row


def _clone_sheet(src, wb):
    """Deep-copy a worksheet (values, styles, merges, widths, hyperlinks) from
    another workbook into `wb`, preserving the title."""
    dst = wb.create_sheet(src.title)
    dst.sheet_state = src.sheet_state
    for col, dim in src.column_dimensions.items():
        d = dst.column_dimensions[col]
        if dim.width:
            d.width = dim.width
        d.hidden = dim.hidden
    for rid, dim in src.row_dimensions.items():
        if dim.height:
            dst.row_dimensions[rid].height = dim.height
    for row in src.iter_rows():
        for c in row:
            nc = dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                nc.font = copy(c.font); nc.fill = copy(c.fill)
                nc.border = copy(c.border); nc.alignment = copy(c.alignment)
                nc.number_format = c.number_format
            if c.hyperlink is not None:
                nc.hyperlink = copy(c.hyperlink)
    for mr in list(src.merged_cells.ranges):
        try:
            dst.merge_cells(str(mr))
        except Exception:
            pass
    if src.freeze_panes:
        dst.freeze_panes = src.freeze_panes
    return dst


def _static_table(ws, df, start_row, green_col=None, bold_total_last=False):
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(no data)").font = Font(name=_FONT, italic=True)
        return start_row + 1
    cols = list(df.columns)
    border = _thin()
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=j, value=str(c))
        cell.font = Font(name=_FONT, bold=True, size=9, color=_HEAD_FG)
        cell.fill = PatternFill("solid", fgColor=_HEAD_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    gidx = cols.index(green_col) + 1 if (green_col and green_col in cols) else None
    for i in range(len(df)):
        r = start_row + 1 + i
        is_total = bold_total_last and i == len(df) - 1
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=j, value=_co(df.iloc[i][c]))
            cell.font = Font(name=_FONT, size=9, bold=is_total)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if not is_total and i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=_ALT)
            if gidx and j == gidx:
                try:
                    fv = float(_co(df.iloc[i][c]))
                    cell.fill = PatternFill("solid", fgColor=_GREEN_BG if fv > 0 else _RED_BG if fv < 0 else "FFFFFF")
                    cell.font = Font(name=_FONT, size=9, bold=is_total,
                                     color=_GREEN_FG if fv > 0 else _RED_FG if fv < 0 else "000000")
                except Exception:
                    pass
    for j, c in enumerate(cols, start=1):
        try:
            w = max([len(str(c))] + [len(str(_co(v))) for v in df[c].head(60)])
        except Exception:
            w = len(str(c))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 9), 30)
    return start_row + 1 + len(df)


# ─────────────────────────────────────────────────────────────────────────────
# Data preparation
# ─────────────────────────────────────────────────────────────────────────────

def _prep(cycle_df, per_trade_df, eligible_ranks_df, candidates_df, leg_rank_df,
          paths, ia, reinvest, cfg=None):
    cfg = cfg or {}
    # ---- per-cycle enrichment + investment merge ----
    inv = ia["window_table"] if ia is not None else pd.DataFrame()
    cyc_rows = []
    for _, r in (cycle_df.iterrows() if cycle_df is not None else []):
        c = int(r["cycle"])
        peak = r.get("peak_close"); buy = r.get("nifty_at_buy")
        trough = r.get("trough_close"); sell = r.get("nifty_at_sell")
        peak_gain = round((peak - buy) / buy * 100, 2) if (peak and buy) else None
        max_dd = round((trough - peak) / peak * 100, 2) if (trough and peak) else None
        recov = round((sell - trough) / trough * 100, 2) if (sell and trough) else None
        tt = per_trade_df[per_trade_df["cycle"] == c] if per_trade_df is not None and not per_trade_df.empty else pd.DataFrame()
        winr = round((tt["return_pct"] > 0).mean() * 100, 1) if len(tt) else None
        best = round(tt["return_pct"].max(), 2) if len(tt) else None
        worst = round(tt["return_pct"].min(), 2) if len(tt) else None
        invr = inv[inv["window"] == c] if not inv.empty else pd.DataFrame()
        cap = float(invr["per_window_capital"].iloc[0]) if not invr.empty else None
        prof = float(invr["window_profit_inr"].iloc[0]) if not invr.empty else None
        eq = float(invr["equity_inr"].iloc[0]) if not invr.empty else None
        cum = float(invr["cumulative_profit_inr"].iloc[0]) if not invr.empty else None
        cyc_rows.append({
            "Cycle": c, "Legs": r.get("legs_used"),
            "WindowStart": _fmt_date(r.get("legs_window_start")),
            "WindowEnd": _fmt_date(r.get("legs_window_end")),
            "BuyDate": _fmt_date(r.get("buy_trigger_date")), "NiftyBuy": buy,
            "PeakClose": peak, "PeakDate": _fmt_date(r.get("peak_date")),
            "PeakGainPct": peak_gain,
            "FallDate": _fmt_date(r.get("fall_confirm_date")), "FallClose": r.get("fall_close"),
            "TroughClose": trough, "TroughDate": _fmt_date(r.get("trough_date")),
            "MaxDDPct": max_dd, "RecoveryPct": recov,
            "SellDate": _fmt_date(r.get("sell_date")), "NiftySell": sell,
            "NiftyCycleRet": r.get("nifty_cycle_return"),
            "NStocks": r.get("n_stocks"), "AvgRet": r.get("avg_return_pct"),
            "WinRatePct": winr, "BestRet": best, "WorstRet": worst,
            "HeldDays": r.get("held_days"),
            "NEffectiveUniverse": r.get("n_effective_universe"),
            "SelectionMode": r.get("selection_mode", "standard"),
            "ExitMode": r.get("exit_mode", "nifty-fall"),
            "NCommon": r.get("n_common"),
            "SelectionFunnel": r.get("selection_funnel"),
            "CapitalINR": round(cap, 2) if cap else None,
            "ProfitINR": round(prof, 2) if prof else None,
            "EquityINR": round(eq, 2) if eq else None,
            "CumProfitINR": round(cum, 2) if cum else None,
            "Status": r.get("status"),
        })
    cyc_tab = pd.DataFrame(cyc_rows)

    # ---- trades table with alloc/profit ----
    tr = per_trade_df.copy() if per_trade_df is not None and not per_trade_df.empty else pd.DataFrame()
    if not tr.empty:
        alloc_map = {}
        if not inv.empty:
            for _, ir in inv.iterrows():
                ns = int(ir.get("n_stocks", 0) or 0)
                alloc_map[int(ir["window"])] = (float(ir["per_window_capital"]) / ns) if ns else 0.0
        tr["AllocINR"] = tr["cycle"].map(lambda c: round(alloc_map.get(int(c), 0.0), 2))
        tr["ProfitINR"] = (tr["AllocINR"] * tr["return_pct"] / 100.0).round(2)
        tr_tab = pd.DataFrame({
            "Cycle": tr["cycle"], "Ticker": tr["ticker"],
            "EntryDate": tr["entry_date"].map(_fmt_date), "EntryPrice": tr["entry_price"],
            "ExitDate": tr["exit_date"].map(_fmt_date), "ExitPrice": tr["exit_price"],
            "ReturnPct": tr["return_pct"], "NiftyRet": tr["nifty_return"], "Alpha": tr["alpha"],
            "DaysHeld": tr["days_held"], "AllocINR": tr["AllocINR"], "ProfitINR": tr["ProfitINR"],
            "Status": tr["status"],
        })
    else:
        tr_tab = pd.DataFrame(columns=["Cycle", "Ticker", "EntryDate", "EntryPrice", "ExitDate",
                                       "ExitPrice", "ReturnPct", "NiftyRet", "Alpha", "DaysHeld",
                                       "AllocINR", "ProfitINR", "Status"])

    # ---- eligible table ----
    el = eligible_ranks_df.copy() if eligible_ranks_df is not None and not eligible_ranks_df.empty else pd.DataFrame()
    if not el.empty:
        el_tab = pd.DataFrame({
            "Cycle": el["cycle"], "Rank": el["rank"], "Ticker": el["ticker"],
            "Quartile": el["quartile"], "EntryDate": el["entry_date"].map(_fmt_date),
            "ExitDate": el["exit_date"].map(_fmt_date), "EntryPrice": el["entry_price"],
            "ExitPrice": el["exit_price"], "ReturnPct": el["return_pct"],
            "NiftyRet": el["nifty_return"], "Alpha": el["alpha"], "NEligible": el["n_eligible"],
            "Traded": el["traded"].map(lambda v: "Y" if bool(v) else "N"),
        })
    else:
        el_tab = pd.DataFrame(columns=["Cycle", "Rank", "Ticker", "Quartile", "EntryDate",
                                       "ExitDate", "EntryPrice", "ExitPrice", "ReturnPct",
                                       "NiftyRet", "Alpha", "NEligible", "Traded"])

    # ---- candidates table (fixed Rise/Fall leg cols) ----
    cd = candidates_df.copy() if candidates_df is not None and not candidates_df.empty else pd.DataFrame()
    if not cd.empty:
        def _legcol(df, suffix):
            for col in df.columns:
                if col.endswith(suffix):
                    return df[col]
            return pd.Series([None] * len(df))
        if cfg.get("metric") == "gate":
            cd_tab = pd.DataFrame({
                "Cycle": cd["cycle"], "CommonRank": cd.get("common_rank"), "Ticker": cd["ticker"],
                "MomentumScore": cd.get("gate_momentum_score") if "gate_momentum_score" in cd.columns else cd.get("mean_roc"),
                "StabilityScore": cd.get("gate_stability_score") if "gate_stability_score" in cd.columns else cd.get("mean_vol"),
                "QualityScore": cd.get("gate_quality_score") if "gate_quality_score" in cd.columns else cd.get("mean_beta_nifty"),
                "PassedMomentum": cd.get("gate_passed_momentum") if "gate_passed_momentum" in cd.columns else "Y",
                "PassedStability": cd.get("gate_passed_stability") if "gate_passed_stability" in cd.columns else "Y",
                "PassedQuality": cd.get("gate_passed_quality") if "gate_passed_quality" in cd.columns else "Y",
                "Selected": cd.get("selected_to_buy").map(lambda v: "Y" if bool(v) else "N") if "selected_to_buy" in cd.columns else "N",
                "Traded": cd.get("traded").map(lambda v: "Y" if bool(v) else "N") if "traded" in cd.columns else "N",
            })
        else:
            cd_tab = pd.DataFrame({
                "Cycle": cd["cycle"], "CommonRank": cd.get("common_rank"), "Ticker": cd["ticker"],
                "MeanMetric": cd.get("mean_metric"),
                "MeanBeta": cd.get("mean_beta_nifty"),
                "BetaRank": cd.get("beta_rank"),
                "MeanCorr": cd.get("mean_corr_nifty"),
                "RiseRank": _legcol(cd, "Rise | rank") if any(c.endswith("Rise | rank") for c in cd.columns) else None,
                "FallRank": _legcol(cd, "Fall | rank") if any(c.endswith("Fall | rank") for c in cd.columns) else None,
                "Selected": cd.get("selected_to_buy").map(lambda v: "Y" if bool(v) else "N") if "selected_to_buy" in cd.columns else "N",
                "Traded": cd.get("traded").map(lambda v: "Y" if bool(v) else "N") if "traded" in cd.columns else "N",
            })
    else:
        cd_tab = pd.DataFrame(columns=["Cycle", "CommonRank", "Ticker", "MeanMetric",
                                       "MeanBeta", "BetaRank", "MeanCorr",
                                       "RiseRank", "FallRank", "Selected", "Traded"])

    # ---- leg-rank table ----
    lr = leg_rank_df.copy() if leg_rank_df is not None and not leg_rank_df.empty else pd.DataFrame()
    if not lr.empty:
        # Normalize per-stock trigger dates → mode date per (cycle/window, leg)
        _grp_cols = [c for c in ["cycle", "window", "leg"] if c in lr.columns]
        if _grp_cols and "entry_date" in lr.columns and "exit_date" in lr.columns:
            _norm_parts = []
            for _, _g in lr.groupby(_grp_cols, sort=False):
                _em = _g["entry_date"].mode()
                _xm = _g["exit_date"].mode()
                if _em.empty or _xm.empty:
                    continue
                _gr = _g.reset_index(drop=True)
                _norm_parts.append(_gr[(_gr["entry_date"] == _em.iloc[0]) & (_gr["exit_date"] == _xm.iloc[0])].copy())
            if _norm_parts:
                lr = pd.concat(_norm_parts).reset_index(drop=True)
        if cfg.get("metric") == "gate":
            lr_tab = pd.DataFrame({
                "Cycle": lr["cycle"], "Leg": lr["leg"],
                "EntryDate": lr["entry_date"].map(_fmt_date), "ExitDate": lr["exit_date"].map(_fmt_date),
                "LegRank": lr["leg_rank"], "Ticker": lr["ticker"],
                "MomentumScore": lr.get("momentum_score"),
                "StabilityScore": lr.get("stability_score"),
                "QualityScore": lr.get("quality_score"),
                "PassedMomentum": lr.get("passed_momentum"),
                "PassedStability": lr.get("passed_stability"),
                "PassedQuality": lr.get("passed_quality"),
                "WindowRank": lr.get("window_rank"), "Common": lr.get("common"), "Bought": lr.get("bought"),
            })
        else:
            lr_tab = pd.DataFrame({
                "Cycle": lr["cycle"], "Leg": lr["leg"],
                "EntryDate": lr["entry_date"].map(_fmt_date), "ExitDate": lr["exit_date"].map(_fmt_date),
                "LegRank": lr["leg_rank"], "Ticker": lr["ticker"],
                "ROC": lr.get("roc_value"), "Volatility": lr.get("vol_value"),
                "Score": lr.get("metric_score") if "metric_score" in lr.columns else lr.get("roc_value"),
                "WindowRank": lr["window_rank"], "Common": lr["common"], "Bought": lr["bought"],
            })
    else:
        lr_tab = pd.DataFrame(columns=["Cycle", "Leg", "EntryDate", "ExitDate", "LegRank",
                                       "Ticker", "ROC", "Volatility", "Score", "WindowRank", "Common", "Bought"])

    # ---- path table + monthwise per cycle ----
    pa = paths.copy() if paths is not None and not paths.empty else pd.DataFrame()
    if not pa.empty:
        pa["_d"] = pd.to_datetime(pa["date"], errors="coerce")
        pa_tab = pd.DataFrame({
            "Cycle": pa["cycle"], "Date": pa["date"].map(_fmt_date),
            "NiftyClose": pa["nifty_close"], "PctFromBuy": pa["pct_from_buy"],
            "RunningPeak": pa["running_peak"], "DDFromPeak": pa["drawdown_from_peak"],
            "Marker": pa["marker"],
        })
        # monthwise within each cycle (NIFTY behaviour by month)
        pm_rows = []
        for c, g in pa.groupby("cycle"):
            g = g.sort_values("_d")
            g["ym"] = g["_d"].dt.to_period("M")
            for ym, gm in g.groupby("ym"):
                o = float(gm["nifty_close"].iloc[0]); cl = float(gm["nifty_close"].iloc[-1])
                pm_rows.append({
                    "Cycle": int(c), "Month": str(ym),
                    "Open": round(o, 2), "Close": round(cl, 2),
                    "MonthRetPct": round((cl - o) / o * 100, 2) if o else None,
                    "High": round(float(gm["nifty_close"].max()), 2),
                    "Low": round(float(gm["nifty_close"].min()), 2),
                    "MaxDDPct": round(float(gm["drawdown_from_peak"].min()), 2),
                    "Days": int(len(gm)),
                })
        pm_tab = pd.DataFrame(pm_rows)
    else:
        pa_tab = pd.DataFrame(columns=["Cycle", "Date", "NiftyClose", "PctFromBuy",
                                       "RunningPeak", "DDFromPeak", "Marker"])
        pm_tab = pd.DataFrame(columns=["Cycle", "Month", "Open", "Close", "MonthRetPct",
                                       "High", "Low", "MaxDDPct", "Days"])

    return cyc_tab, tr_tab, el_tab, cd_tab, lr_tab, pa_tab, pm_tab


def _overall_time_summary(per_trade_df, freq="Y"):
    """Year/month-wise summary across all trades (by entry date)."""
    if per_trade_df is None or per_trade_df.empty:
        return pd.DataFrame()
    df = per_trade_df.copy()
    df["_d"] = pd.to_datetime(df["entry_date"], errors="coerce")
    df["key"] = df["_d"].dt.year.astype("Int64").astype(str) if freq == "Y" \
        else df["_d"].dt.to_period("M").astype(str)
    rows = []
    for k, g in df.groupby("key"):
        ret = pd.to_numeric(g["return_pct"], errors="coerce").dropna()
        alp = pd.to_numeric(g.get("alpha"), errors="coerce").dropna()
        nif = pd.to_numeric(g.get("nifty_return"), errors="coerce").dropna()
        rows.append({
            "Period": k, "Cycles": int(g["cycle"].nunique()), "Trades": int(len(g)),
            "GrossPct": round(float(ret.sum()), 2) if len(ret) else 0,
            "AvgPct": round(float(ret.mean()), 2) if len(ret) else 0,
            "WinPct": round(float((ret > 0).mean() * 100), 1) if len(ret) else 0,
            "AvgAlpha": round(float(alp.mean()), 2) if len(alp) else 0,
            "AvgNifty": round(float(nif.mean()), 2) if len(nif) else 0,
            "BestPct": round(float(ret.max()), 2) if len(ret) else 0,
            "WorstPct": round(float(ret.min()), 2) if len(ret) else 0,
        })
    out = pd.DataFrame(rows).sort_values("Period").reset_index(drop=True)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def generate_momentum_interactive_excel(
    per_trade_df, cycle_df, candidates_df, status_df, phases, config, *,
    eligible_ranks_df=None, leg_rank_df=None, audit_df=None, vol_audit_df=None,
    nifty_df=None, stock_dict=None,
    returns_df=None, initial_capital=100_000.0, reinvest=True,
):
    cfg = config or {}
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    ia = None
    if per_trade_df is not None and not per_trade_df.empty \
            and per_trade_df["window_idx"].nunique() >= 1:
        try:
            ia = compute_investment_analysis(per_trade_df, initial_capital=initial_capital,
                                             alloc_mode="equal", reinvest=reinvest)
        except Exception:
            ia = None

    # NIFTY paths (day-by-day per cycle)
    paths = pd.DataFrame()
    if cycle_df is not None and not cycle_df.empty and nifty_df is not None and not nifty_df.empty:
        rows = []
        for _, c in cycle_df.iterrows():
            b = pd.Timestamp(c["buy_trigger_date"]); s = pd.Timestamp(c["sell_date"])
            seg = nifty_df[(nifty_df.index >= b) & (nifty_df.index <= s)]
            if seg.empty:
                continue
            buy_px = float(seg["close"].iloc[0]); peak = -1e18
            pkd = pd.Timestamp(c["peak_date"]) if pd.notna(c.get("peak_date")) else None
            fld = pd.Timestamp(c["fall_confirm_date"]) if pd.notna(c.get("fall_confirm_date")) else None
            for dt, px in seg["close"].items():
                px = float(px); peak = max(peak, px)
                mk = "BUY" if dt == b else ("PEAK" if pkd is not None and dt == pkd else
                     ("FALL" if fld is not None and dt == fld else ("SELL" if dt == s else "")))
                rows.append({"cycle": int(c["cycle"]), "date": dt, "nifty_close": round(px, 2),
                             "pct_from_buy": round((px - buy_px) / buy_px * 100, 2) if buy_px else 0,
                             "running_peak": round(peak, 2),
                             "drawdown_from_peak": round((px - peak) / peak * 100, 2) if peak > 0 else 0,
                             "marker": mk})
        paths = pd.DataFrame(rows)

    cyc_tab, tr_tab, el_tab, cd_tab, lr_tab, pa_tab, pm_tab = _prep(
        cycle_df, per_trade_df, eligible_ranks_df, candidates_df, leg_rank_df, paths, ia, reinvest, cfg)

    cycles = [int(x) for x in cyc_tab["Cycle"].tolist()] if not cyc_tab.empty else []
    default_cycle = cycles[0] if cycles else 1

    # ── Hidden DATA tables ─────────────────────────────────────────────────────
    _df_to_table(wb, "_CycleData", cyc_tab, "tblCycles")
    _df_to_table(wb, "_TradesData", tr_tab, "tblTrades")
    _df_to_table(wb, "_EligData", el_tab, "tblElig")
    _df_to_table(wb, "_CandData", cd_tab, "tblCand")
    _df_to_table(wb, "_LegData", lr_tab, "tblLeg")
    _df_to_table(wb, "_PathData", pa_tab, "tblPath")
    _df_to_table(wb, "_MonthData", pm_tab, "tblMonth")

    CTRL = "Dashboard!$C$2"

    # ══ ① DASHBOARD ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard"
    # Build a customised title from the active momentum settings
    _m = cfg.get("metric", "roc")
    _m_lbl = (
        "Off (Vol-filter)" if _m == "off"
        else "ROC" if _m == "roc"
        else f"Volatility ({cfg.get('vol_dir','low')})" if _m == "vol"
        else "Vol×ROC" if _m == "volxroc"
        else "ROC/Vol" if _m == "roc_over_vol"
        else "Beta/Vol (β÷vol)" if _m == "beta_over_vol"
        else "Beta×Vol (β×vol)" if _m == "beta_x_vol"
        else "Beta-NIFTY (lowest β)" if _m == "beta"
        else "Std Dev / Downside Vol (σ÷DV)" if _m == "sd_over_dv"
        else f"Both-ROC∩Vol({cfg.get('vol_dir','low')})"
    )
    _k     = cfg.get("top_k", "?")
    _np    = cfg.get("nifty_pct", "?")
    _fp    = cfg.get("fall", "?")
    _hold  = cfg.get("max_hold", "?")
    _side  = "Top" if cfg.get("selection_side","top") == "top" else "Bottom"
    _vf    = cfg.get("vol_filter", "off")
    _vf_sfx = f"  ·  VF-{_vf}" if _vf != "off" else ""
    if _m == "off":
        _topn_sfx = f"  ·  Top-{cfg.get('vol_filter_n','?')} per window"
        _side_sfx = ""
    else:
        _topn_sfx = f"  ·  " + ("Bot" if cfg.get("selection_side","top") == "bottom" else "Top") + f"-N={cfg.get('top_n', cfg.get('topn_rise','?'))}"
        _side_sfx = f"  ·  {_side}-N"
    _run_name = _make_run_name(cfg)
    _dash_title = f"S³ — {_run_name}"
    _title(ws, _dash_title, 8)
    ws.cell(row=2, column=1, value="▶ Type a CYCLE # →").font = Font(name=_FONT, bold=True, size=12)
    ctrl = ws.cell(row=2, column=3, value=default_cycle)
    ctrl.font = Font(name=_FONT, bold=True, size=14, color="9C3B00")
    ctrl.fill = PatternFill("solid", fgColor=_CTRL)
    ctrl.alignment = Alignment(horizontal="center", vertical="center")
    ctrl.border = _thin()
    ws.cell(row=2, column=4,
            value=f"(valid 1–{max(cycles) if cycles else 1}; every View sheet auto-updates)").font = \
        Font(name=_FONT, italic=True, size=9, color="666666")
    if cycles:
        dv = DataValidation(type="whole", operator="between",
                            formula1=str(min(cycles)), formula2=str(max(cycles)),
                            allow_blank=False)
        dv.error = "Enter a valid cycle number."
        dv.prompt = "Type a cycle number to drill in."
        ws.add_data_validation(dv); dv.add(ctrl)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    def mi(col):  # INDEX/MATCH against tblCycles for the selected cycle
        return f'=IFERROR(INDEX(tblCycles[{col}],MATCH({CTRL},tblCycles[Cycle],0)),"—")'

    r = 4
    _band(ws, r, "Selected Cycle — Summary", 8); r += 1
    summ = [
        ("Legs used", mi("Legs")), ("Window start", mi("WindowStart")),
        ("Window end", mi("WindowEnd")), ("Buy date", mi("BuyDate")),
        ("NIFTY at buy", mi("NiftyBuy")), ("Peak close", mi("PeakClose")),
        ("Peak date", mi("PeakDate")), ("Peak gain %", mi("PeakGainPct")),
        ("Fall confirm date", mi("FallDate")), ("Trough close", mi("TroughClose")),
        ("Trough date", mi("TroughDate")), ("Max drawdown %", mi("MaxDDPct")),
        ("Recovery %", mi("RecoveryPct")), ("Sell date", mi("SellDate")),
        ("NIFTY at sell", mi("NiftySell")), ("NIFTY cycle return %", mi("NiftyCycleRet")),
        ("Held days", mi("HeldDays")), ("Status", mi("Status")),
        ("Effective universe (selection + vol filter)", mi("NEffectiveUniverse")),
        ("Common across legs", mi("NCommon")),
        ("Selection funnel", mi("SelectionFunnel")),
    ]
    for k, f in summ:
        ws.cell(row=r, column=1, value=k).font = Font(name=_FONT, bold=True, size=10)
        ws.cell(row=r, column=3, value=f).font = Font(name=_FONT, size=10)
        r += 1

    r += 1
    _band(ws, r, "Selected Cycle — Investment & Trade Stats", 8); r += 1
    stats = [
        ("Stocks held", f"=COUNTIFS(tblTrades[Cycle],{CTRL})"),
        ("Capital deployed (₹)", mi("CapitalINR")),
        ("Cycle profit (₹)", mi("ProfitINR")),
        ("Equity after cycle (₹)", mi("EquityINR")),
        ("Cumulative profit (₹)", mi("CumProfitINR")),
        ("Avg stock return %", f'=IFERROR(AVERAGEIFS(tblTrades[ReturnPct],tblTrades[Cycle],{CTRL}),"—")'),
        ("Win rate %", f'=IFERROR(COUNTIFS(tblTrades[Cycle],{CTRL},tblTrades[ReturnPct],">0")/COUNTIFS(tblTrades[Cycle],{CTRL})*100,"—")'),
        ("Avg alpha %", f'=IFERROR(AVERAGEIFS(tblTrades[Alpha],tblTrades[Cycle],{CTRL}),"—")'),
        ("Best stock return %", mi("BestRet")),
        ("Worst stock return %", mi("WorstRet")),
        ("Avg days held", f'=IFERROR(AVERAGEIFS(tblTrades[DaysHeld],tblTrades[Cycle],{CTRL}),"—")'),
    ]
    for k, f in stats:
        ws.cell(row=r, column=1, value=k).font = Font(name=_FONT, bold=True, size=10)
        ws.cell(row=r, column=3, value=f).font = Font(name=_FONT, size=10)
        r += 1

    r += 1
    _band(ws, r, "Selected Cycle — Bought-stock Quartile Distribution", 8); r += 1
    for q in ["Q4", "Q3", "Q2", "Q1"]:
        ws.cell(row=r, column=1, value=f"{q} bought").font = Font(name=_FONT, bold=True, size=10)
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS(tblElig[Cycle],{CTRL},tblElig[Traded],"Y",tblElig[Quartile],"{q}")') \
            .font = Font(name=_FONT, size=10)
        r += 1
    ws.cell(row=r, column=1, value="Eligible (NIFTY 500 universe on buy)").font = Font(name=_FONT, bold=True, size=10)
    ws.cell(row=r, column=3, value=f"=COUNTIFS(tblElig[Cycle],{CTRL})").font = Font(name=_FONT, size=10)
    r += 2
    ws.cell(row=r, column=1,
            value="Open the View sheets (Trades / Eligible / Candidates / Leg Rankings / "
                  "NIFTY Path / Monthwise) — they all follow the cycle # above.") \
        .font = Font(name=_FONT, italic=True, size=9, color="666666")

    # max rows any single cycle can have, per table (for the formula grid height)
    def _kmax(df):
        if df is None or df.empty or "Cycle" not in df.columns:
            return 1
        return int(df.groupby("Cycle").size().max())
    kmax_tr = _kmax(tr_tab); kmax_el = _kmax(el_tab); kmax_cd = _kmax(cd_tab)
    kmax_lr = _kmax(lr_tab); kmax_pa = _kmax(pa_tab); kmax_pm = _kmax(pm_tab)

    # ══ View sheets (auto-filter by the Dashboard cycle #) ═══════════════════════
    v = wb.create_sheet("View · Trades")
    _title(v, "Cycle Trades — auto-filtered by Dashboard cycle #", 12)
    _filter_view(v, 3,
                 ["Ticker", "Entry Date", "Entry Price", "Exit Date", "Exit Price",
                  "Return %", "Nifty Ret %", "Alpha %", "Days Held", "Alloc ₹", "Profit ₹", "Status"],
                 "tblTrades",
                 ["Ticker", "EntryDate", "EntryPrice", "ExitDate", "ExitPrice",
                  "ReturnPct", "NiftyRet", "Alpha", "DaysHeld", "AllocINR", "ProfitINR", "Status"],
                 kmax_tr, CTRL, green_col="Return %")

    v = wb.create_sheet("View · Eligible Ranking")
    _title(v, "Eligible Stock Ranking (NIFTY 500 Universe) — auto-filtered · Q4..Q1 · Traded Y/N", 13)
    _filter_view(v, 3,
                 ["Rank", "Ticker", "Quartile", "Entry Date", "Exit Date", "Entry Price",
                  "Exit Price", "Return %", "Nifty Ret %", "Alpha %", "N Eligible", "Traded"],
                 "tblElig",
                 ["Rank", "Ticker", "Quartile", "EntryDate", "ExitDate", "EntryPrice",
                  "ExitPrice", "ReturnPct", "NiftyRet", "Alpha", "NEligible", "Traded"],
                 kmax_el, CTRL, green_col="Return %", traded_col="Traded")

    v = wb.create_sheet("View · Candidates")
    if cfg.get("metric") == "gate":
        _title(v, "Cycle Candidates — Sequential Gate Filtration (Momentum → Stability → Quality) — auto-filtered", 10)
        _filter_view(v, 3,
                     ["Common Rank", "Ticker", "Momentum Score", "Stability Score", "Quality Score",
                      "Passed Momentum", "Passed Stability", "Passed Quality", "Selected", "Traded"],
                     "tblCand",
                     ["CommonRank", "Ticker", "MomentumScore", "StabilityScore", "QualityScore",
                      "PassedMomentum", "PassedStability", "PassedQuality", "Selected", "Traded"],
                     kmax_cd, CTRL, traded_col="Traded")
    else:
        _cand_title = ("Cycle Candidates — Common ranking by Beta (ascending) · Mean Beta · Beta Rank — auto-filtered"
                       if cfg.get("metric") == "beta"
                       else "Cycle Candidates — Common ranking by Beta/Vol (β÷vol) — auto-filtered"
                       if cfg.get("metric") == "beta_over_vol"
                       else "Cycle Candidates — Common ranking by Beta×Vol (β×vol) — auto-filtered"
                       if cfg.get("metric") == "beta_x_vol"
                       else "Cycle Candidates — Common ranking by Std Dev / Downside Vol (σ÷DV) — auto-filtered"
                       if cfg.get("metric") == "sd_over_dv"
                       else "Cycle Candidates (common ranking) — auto-filtered")
        _title(v, _cand_title, 10)
        _filter_view(v, 3,
                     ["Common Rank", "Ticker", "Mean Metric", "Mean Beta", "Beta Rank", "Mean Corr",
                      "Rise Rank", "Fall Rank", "Selected", "Traded"],
                     "tblCand",
                     ["CommonRank", "Ticker", "MeanMetric", "MeanBeta", "BetaRank", "MeanCorr",
                      "RiseRank", "FallRank", "Selected", "Traded"],
                     kmax_cd, CTRL, traded_col="Traded")

    v = wb.create_sheet("View · Leg Rankings")
    _side_lbl = "Bottom-N" if cfg.get("selection_side", "top") == "bottom" else "Top-N"
    _cfg_metric = cfg.get("metric", "roc")
    _cfg_vt = cfg.get("vol_type", "standard")
    
    if _cfg_metric == "gate":
        _title(v, "Cycle Leg Rankings (Gate System Scorecards & Sequential Filtration) — auto-filtered", 12)
        _filter_view(v, 3,
                     ["Leg", "Entry Date", "Exit Date", "Leg Rank", "Ticker", "Momentum Score",
                      "Stability Score", "Quality Score", "Passed Momentum", "Passed Stability",
                      "Passed Quality", "Bought"],
                     "tblLeg",
                     ["Leg", "EntryDate", "ExitDate", "LegRank", "Ticker", "MomentumScore",
                      "StabilityScore", "QualityScore", "PassedMomentum", "PassedStability",
                      "PassedQuality", "Bought"],
                     kmax_lr, CTRL, traded_col="Bought")
    else:
        _lr_vol_hdr = "Downside Vol" if _cfg_vt == "downside" else "Volatility"
        _lr_score_hdr = {
            "roc_over_vol": "ROC/Vol Score",
            "volxroc": "ROC×Vol Score",
            "vol": "Vol Score",
            "roc": "ROC Score",
            "both": "ROC Score",
            "off": "Vol Score",
            "beta": "Beta (NIFTY)",
            "beta_over_vol": "Beta/Vol Score",
            "beta_x_vol": "Beta×Vol Score",
        }.get(_cfg_metric, "Score")
        if _cfg_metric in ("beta", "beta_over_vol", "beta_x_vol"):
            _bov_sfx = (" (β÷vol)" if _cfg_metric == "beta_over_vol"
                        else " (β×vol)" if _cfg_metric == "beta_x_vol"
                        else "")
            _leg_rank_title = (f"Cycle Leg Rankings — {_side_lbl} by Beta{_bov_sfx} (NIFTY) "
                               f"Rank 1 = lowest beta · base-day prepended — auto-filtered")
        elif cfg.get("metric") == "off":
            _leg_rank_title = f"Volatility Window Rankings ({_side_lbl} per 100-day window) — auto-filtered"
        else:
            _leg_rank_title = f"Cycle Leg Rankings ({_side_lbl} per leg) — auto-filtered"
        _title(v, _leg_rank_title, 10)
        _filter_view(v, 3,
                     ["Leg", "Entry Date", "Exit Date", "Leg Rank", "Ticker", "ROC %", _lr_vol_hdr,
                      _lr_score_hdr, "Window Rank", "Common", "Bought"],
                     "tblLeg",
                     ["Leg", "EntryDate", "ExitDate", "LegRank", "Ticker", "ROC", "Volatility",
                      "Score", "WindowRank", "Common", "Bought"],
                     kmax_lr, CTRL, traded_col="Bought")

    v = wb.create_sheet("View · NIFTY Path")
    _title(v, "Cycle NIFTY Path (day-by-day) — auto-filtered", 7)
    _filter_view(v, 3,
                 ["Date", "Nifty Close", "% From Buy", "Running Peak", "DD From Peak", "Marker"],
                 "tblPath",
                 ["Date", "NiftyClose", "PctFromBuy", "RunningPeak", "DDFromPeak", "Marker"],
                 kmax_pa, CTRL, green_col="% From Buy")

    v = wb.create_sheet("View · Monthwise (cycle)")
    _title(v, "Selected Cycle — Month-by-Month NIFTY behaviour — auto-filtered", 9)
    _filter_view(v, 3,
                 ["Month", "Open", "Close", "Month Ret %", "High", "Low", "Max DD %", "Days"],
                 "tblMonth",
                 ["Month", "Open", "Close", "MonthRetPct", "High", "Low", "MaxDDPct", "Days"],
                 kmax_pm, CTRL, green_col="Month Ret %")

    # ══ Cycle Ledger (all) — quartile-enriched, NIFTY prices removed ════════════
    # Per-cycle quartile counts of BOUGHT stocks (from the eligible ranking).
    qmap = {}
    if eligible_ranks_df is not None and not eligible_ranks_df.empty:
        for c, g in eligible_ranks_df[eligible_ranks_df["traded"] == True].groupby("cycle"):  # noqa: E712
            qc = g["quartile"].value_counts().to_dict()
            q1, q2, q3, q4 = (int(qc.get(f"Q{i}", 0)) for i in (1, 2, 3, 4))
            tot = q1 + q2 + q3 + q4
            qmap[int(c)] = (q1, q2, q3, q4, q3 + q4,
                            round((q3 + q4) / tot * 100, 1) if tot else 0.0)

    led = cyc_tab.copy()
    drop_price = ["NiftyBuy", "PeakClose", "FallClose", "TroughClose", "NiftySell"]
    led = led.drop(columns=[c for c in drop_price if c in led.columns])
    led["Q1"] = led["Cycle"].map(lambda c: qmap.get(int(c), (0,)*6)[0])
    led["Q2"] = led["Cycle"].map(lambda c: qmap.get(int(c), (0,)*6)[1])
    led["Q3"] = led["Cycle"].map(lambda c: qmap.get(int(c), (0,)*6)[2])
    led["Q4"] = led["Cycle"].map(lambda c: qmap.get(int(c), (0,)*6)[3])
    led["Q3+Q4"] = led["Cycle"].map(lambda c: qmap.get(int(c), (0,)*6)[4])
    led["Q3+Q4 %"] = led["Cycle"].map(lambda c: qmap.get(int(c), (0,)*6)[5])
    # tidy order: identity → dates/% → quartiles → stats → portfolio
    order = [c for c in ["Cycle", "Legs", "WindowStart", "WindowEnd", "BuyDate",
             "PeakDate", "PeakGainPct", "FallDate", "TroughDate", "MaxDDPct",
             "RecoveryPct", "SellDate", "NiftyCycleRet", "NStocks",
             "NEffectiveUniverse", "NCommon",
             "Q1", "Q2", "Q3", "Q4", "Q3+Q4", "Q3+Q4 %",
             "AvgRet", "WinRatePct", "BestRet", "WorstRet", "HeldDays",
             "CapitalINR", "ProfitINR", "EquityINR", "CumProfitINR",
             "SelectionFunnel", "Status"]
             if c in led.columns]
    led = led[order]

    wl = wb.create_sheet("Cycle Ledger (all)")
    _title(wl, "Cycle Ledger — every cycle · per-cycle Q1–Q4 / Q3+Q4 / Q3+Q4 % of bought "
           "stocks · click ▶ to open a cycle's detail", len(led.columns) + 1)
    led.insert(0, "Open", ["▶ Detail"] * len(led))
    _static_table(wl, led, 3, green_col="AvgRet")

    # ══ Overall Monthwise (yearwise is provided by the detailed 'Yearwise Summary') ═
    wm = wb.create_sheet("Overall Monthwise")
    _title(wm, "Overall Month-wise Summary (all cycles, by trade entry month)", 10)
    _static_table(wm, _overall_time_summary(per_trade_df, "M"), 3, green_col="GrossPct")

    # ══ Calculation Audit (Fix #8) — schema adapts to the setting ═══════════════
    wca = wb.create_sheet("Calculation Audit")
    if audit_df is not None and not audit_df.empty and "roc" in audit_df.columns:
        # ROC / ROC×Vol / Both / Vol ranking schema
        ad = audit_df.copy()
        for dc in ["leg_entry_date", "leg_exit_date", "fall_entry_date", "fall_exit_date",
                   "roc_start_date", "roc_end_date", "vol_start_date", "vol_end_date"]:
            if dc in ad.columns:
                ad[dc] = ad[dc].map(_fmt_date)
        _metric = cfg.get("metric", "roc")
        _vt = cfg.get("vol_type", "standard")
        _vol_lbl = "Downside Vol" if _vt == "downside" else "Volatility"
        _score_lbl = {
            "roc_over_vol": "ROC / Vol (Score)",
            "volxroc":      "ROC × Vol (Score)",
            "vol":          "Vol (Score)",
            "roc":          "ROC (Score)",
            "both":         "ROC (Score)",
            "beta":         "Beta (NIFTY ln-ret)",
            "beta_over_vol": "Beta/Vol (Score)",
            "beta_x_vol":   "Beta×Vol (Score)",
            "sd_over_dv":   "Std/Downside Ratio",
        }.get(_metric, "Score")
        order = [c for c in [
            "cycle", "ticker", "leg", "trade",
            "leg_entry_date", "leg_exit_date", "fall_entry_date", "fall_exit_date",
            "roc_start_date", "roc_end_date", "roc",
            "vol_start_date", "vol_end_date", "vol_trading_days", "volatility",
            "roc_x_vol", "roc_over_vol", "beta_over_vol", "beta_x_vol",
            "sd_std_vol", "sd_dv_vol", "sd_over_dv", "metric_score",
            "beta_nifty", "corr_nifty", "beta_rank",
            "qualified_for_ranking", "in_leg_topn", "selected_to_buy",
            "reason"] if c in ad.columns]
        ad = ad[order].rename(columns={
            "cycle": "Cycle", "ticker": "Stock", "leg": "Leg", "trade": "Trade",
            "leg_entry_date": "Window Start (Actual)", "leg_exit_date": "Window End (Actual)",
            "fall_entry_date": "Fall Entry", "fall_exit_date": "Fall Exit",
            "roc_start_date": "ROC Start Date", "roc_end_date": "ROC End Date",
            "roc": "ROC %", "vol_start_date": "Vol Window Start",
            "vol_end_date": "Vol Window End", "vol_trading_days": "Vol Trading Days",
            "volatility": _vol_lbl, "roc_x_vol": "ROC × Vol",
            "roc_over_vol": "ROC / Vol",
            "beta_over_vol": "Beta / Vol",
            "beta_x_vol": "Beta × Vol",
            "sd_std_vol": "Std Volatility",
            "sd_dv_vol":  "Downside Volatility",
            "sd_over_dv": "Std/Downside Ratio",
            "metric_score": _score_lbl,
            "beta_nifty": "Beta (NIFTY)", "corr_nifty": "Corr (NIFTY)", "beta_rank": "Beta Rank",
            "qualified_for_ranking": "Qualified For Ranking",
            "in_leg_topn": ("In Leg Bot-N" if cfg.get("selection_side","top") == "bottom" else "In Leg Top-N"), "selected_to_buy": "Selected?",
            "reason": "Reason"})
        _title(wca, "Calculation Audit — every (cycle, stock, leg) decision: "
                    "ROC window, volatility window, qualification & reason",
               len(ad.columns))
        _static_table(wca, ad, 3)
    elif audit_df is not None and not audit_df.empty and "common" in audit_df.columns:
        # OFF (pure 100-day volatility) schema — the selection basis per stock
        ad = audit_df.copy()
        for dc in [c for c in ad.columns if c.endswith(("_start", "_end"))
                   or c in ("fall_entry_date", "fall_exit_date")]:
            ad[dc] = ad[dc].map(_fmt_date)
        # dynamic window columns w1_*, w2_* …
        wcols = []
        i = 1
        while f"w{i}_vol" in ad.columns:
            base = [f"w{i}_ref", f"w{i}_vol", f"w{i}_start", f"w{i}_end",
                    f"w{i}_days", f"w{i}_rank", f"w{i}_in_topn"]
            wcols += [c for c in base if c in ad.columns]
            i += 1
            if i > 6:
                break
        n_win = i - 1
        order = (["cycle", "ticker", "fall_entry_date", "fall_exit_date"] + wcols
                 + ["common", "selected_to_buy", "reason"])
        order = [c for c in order if c in ad.columns]
        ren = {"cycle": "Cycle", "ticker": "Stock",
               "fall_entry_date": "Fall Entry", "fall_exit_date": "Fall Exit",
               "common": "Common (all windows)?", "selected_to_buy": "Bought?",
               "reason": "Reason"}
        for k in range(1, n_win + 1):
            ren[f"w{k}_ref"] = f"W{k} Window"
            ren[f"w{k}_vol"] = f"W{k} 100-Day Vol"
            ren[f"w{k}_start"] = f"W{k} Vol Start"
            ren[f"w{k}_end"] = f"W{k} Vol End"
            ren[f"w{k}_days"] = f"W{k} Trading Days"
            ren[f"w{k}_rank"] = f"W{k} Vol Rank"
            ren[f"w{k}_in_topn"] = f"W{k} In " + ("Bot-N" if cfg.get("selection_side","top") == "bottom" else "Top-N")
        ad = ad[order].rename(columns=ren)
        _title(wca, "Calculation Audit — 100-day volatility selection basis: each "
                    "stock's volatility, dates & rank in every window, whether common "
                    "to all windows, and whether it was bought", len(ad.columns))
        _static_table(wca, ad, 3)
    else:
        _title(wca, "Calculation Audit", 2)
        wca.cell(row=3, column=1, value="(no selection audit available for this run)")

    # ══ 100-Day Volatility Audit (Fix #7) ═══════════════════════════════════════
    wva = wb.create_sheet("100-Day Volatility Audit")
    if vol_audit_df is not None and not vol_audit_df.empty:
        va = vol_audit_df.copy()
        for dc in ["as_of_date", "vol_window_start", "vol_window_end"]:
            if dc in va.columns:
                va[dc] = va[dc].map(_fmt_date)
        order = [c for c in [
            "cycle", "ticker", "reference", "as_of_date", "requested_trading_days",
            "vol_window_start", "vol_window_end", "trading_days_used",
            "volatility", "in_top_n"] if c in va.columns]
        va = va[order].rename(columns={
            "cycle": "Cycle", "ticker": "Stock", "reference": "Reference Window",
            "as_of_date": "As-Of Date", "requested_trading_days": "Requested Trading Days",
            "vol_window_start": "Volatility Start Date", "vol_window_end": "Volatility End Date",
            "trading_days_used": "Trading Days Used", "volatility": "100-Day Volatility",
            "in_top_n": ("In Bot-N" if cfg.get("selection_side","top") == "bottom" else "In Top-N")})
        _title(wva, "100-Day Volatility Audit — exact window start/end dates and trading "
                    "days used for every volatility computation", len(va.columns))
        _static_table(wva, va, 3)
    else:
        _title(wva, "100-Day Volatility Audit", 2)
        wva.cell(row=3, column=1, value="(volatility filter Off — no point-in-time volatility windows)")
    if cfg.get("gate_params"):
        wgp = wb.create_sheet("Gate Parameters")
        _title(wgp, "Gate System Parameters", 3)
        from core.gate_system import params_summary
        g_df = params_summary(cfg["gate_params"])
        _static_table(wgp, g_df, 3)
        wgp.column_dimensions["A"].width = 25
        wgp.column_dimensions["B"].width = 30
        wgp.column_dimensions["C"].width = 40

    # ══ MERGE: bring in EVERY sheet from the detailed workbook ══════════════════
    detail_cycle_sheets = set()
    try:
        det_bytes = generate_momentum_excel(
            per_trade_df=per_trade_df, cycle_df=cycle_df, candidates_df=candidates_df,
            status_df=status_df, phases=phases, config=cfg,
            eligible_ranks_df=eligible_ranks_df, leg_rank_df=leg_rank_df,
            nifty_df=nifty_df, stock_dict=stock_dict, returns_df=returns_df,
            initial_capital=initial_capital, reinvest=reinvest)
        det_wb = load_workbook(io.BytesIO(det_bytes))
        existing = set(wb.sheetnames)
        # the detailed 'Cycle Ledger' is replaced by our enriched 'Cycle Ledger (all)'
        skip = {"Cycle Ledger"}
        for sn in det_wb.sheetnames:
            if sn in skip or sn in existing:
                continue
            _clone_sheet(det_wb[sn], wb)
            if sn.endswith(" Detail"):
                detail_cycle_sheets.add(sn)
    except Exception:
        pass

    # ══ Hyperlink the enriched ledger's ▶ cells to each per-cycle Detail sheet ═══
    if not led.empty:
        for i, cval in enumerate(led["Cycle"].tolist()):
            sn = f"Cycle {int(cval)} Detail"
            if sn in detail_cycle_sheets:
                cell = wl.cell(row=4 + i, column=1)
                cell.value = "▶ Detail"
                cell.hyperlink = f"#'{sn}'!A1"
                cell.font = Font(name=_FONT, color="0563C1", underline="single", bold=True, size=9)

    # order: Dashboard first; push hidden data sheets to the end
    wb.move_sheet("Dashboard", -(wb.sheetnames.index("Dashboard")))
    for hidden in ["_CycleData", "_TradesData", "_EligData", "_CandData",
                   "_LegData", "_PathData", "_MonthData"]:
        if hidden in wb.sheetnames:
            wb.move_sheet(hidden, len(wb.sheetnames) - wb.sheetnames.index(hidden) - 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
