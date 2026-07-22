"""
S³ Export — Momentum Portfolio Report (detailed)
================================================
Full, per-window/per-cycle Excel workbook for the Momentum Based Investment
strategy — mirrors the depth of the short-term exporter:

  ①  Portfolio Summary    — config + headline metrics + portfolio stats + period
  ②  Cycle Ledger         — buy → peak → fall → trough → sell, NIFTY levels
  ③  Cycle Status         — EVERY window: traded / skipped + reason
  ④  Cycle Candidates     — per-cycle pure-ranking common stocks (per-leg detail)
  ⑤  Window Leg Rankings  — per-window Top-N per leg (ROC/Vol), every window
  ⑥  Per-Cycle Equity     — capital / return / profit / equity per cycle
  ⑦  Common Stocks P&L    — each traded stock: buy, sell, return, alpha, P&L
  ⑧  Trade Log            — every executed trade
  ⑨  Stock Summary        — per-ticker aggregated statistics
  ⑩  NIFTY Cycle Path     — day-by-day NIFTY per cycle (peak / drawdown / markers)
  ⑪  Yearwise Summary     — per calendar year + TOTAL
  ⑫  Phase Schedule       — full phase listing
"""
from __future__ import annotations

import io

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from core.investment_analysis import compute_investment_analysis
    from core.multi_leg_engine import find_pattern_windows, _window_vol
    from core.momentum_engine import _leg_roc_map, _leg_vol_map
    from core.phase_engine import quartile_label
except Exception:  # pragma: no cover
    from investment_analysis import compute_investment_analysis  # type: ignore
    from multi_leg_engine import find_pattern_windows, _window_vol  # type: ignore
    from momentum_engine import _leg_roc_map, _leg_vol_map  # type: ignore
    from phase_engine import quartile_label  # type: ignore

_FONT      = "Arial"
_HEADER_BG = "2E4057"
_HEADER_FG = "FFFFFF"
_TITLE_BG  = "6C63FF"
_SUB_BG    = "44476A"

def _make_run_name(cfg: dict) -> str:
    """Compact settings-based run name used for sheet titles and filenames."""
    m   = cfg.get("metric", "roc")
    vt  = cfg.get("vol_type", "standard")
    vd  = cfg.get("vol_dir", cfg.get("vol_direction", "low"))
    side = cfg.get("selection_side", "top")
    top_n = cfg.get("top_n", cfg.get("vol_filter_n", 50))
    top_k = cfg.get("top_k", 10)
    use_ff = cfg.get("use_fixed_fall", False)
    fe_d  = cfg.get("fall_entry_days", 100)
    fx_d  = cfg.get("fall_exit_days", 100)
    np_   = cfg.get("nifty_pct", 5)
    fp_   = cfg.get("fall_pct", 10)
    hd_   = cfg.get("max_hold", 2)
    exact = cfg.get("exact_trigger_mode", False)
    _ds = "DS" if vt == "downside" else "Std"
    _lo = "Lo" if vd == "low" else "Hi"
    m_tok = {
        "roc":         "ROC",
        "vol":         f"Vol-{_ds}-{_lo}",
        "volxroc":     f"RxV-{_ds}",
        "roc_over_vol":f"R/V-{_ds}-{_lo}",
        "both":        f"Both-{_ds}-{_lo}",
        "beta":        "Beta",
        "beta_over_vol": f"BetaOVol-{_ds}",
        "beta_x_vol":  f"BetaXVol-{_ds}",
        "sd_over_dv":  "SdOvDV",
        "gate":        "Gate-ARQM",
        "off":         f"Off-{_ds}-{_lo}",
    }.get(m, m.upper())
    win_tok  = f"FF{fe_d}+{fx_d}d" if use_ff else "AutoLegs"
    side_tok = "Bot" if side == "bottom" else "Top"
    exact_tok = " ExactTrig" if exact else ""
    return (f"{m_tok} {side_tok}{top_n} K{top_k} {win_tok} "
            f"N{np_:g}% Fall{fp_:g}% Hold{hd_:g}yr{exact_tok}")

_GREEN_BG  = "C6EFCE"; _GREEN_FG = "1A6B3C"
_RED_BG    = "FFC7CE"; _RED_FG   = "9C0006"
_GOLD_BG   = "FFF2CC"
_ALT       = "F2F4FF"
_BORDER    = "B8C2D9"


def _thin():
    s = Side(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_date(d) -> str:
    if d is None or (isinstance(d, float) and (np.isnan(d) if isinstance(d, float) else False)):
        return ""
    try:
        if pd.isna(d):
            return ""
    except Exception:
        pass
    try:
        return pd.Timestamp(d).strftime("%d-%b-%Y")
    except Exception:
        return str(d)


def _title(ws, text, ncols, bg=_TITLE_BG):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, ncols))
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=_FONT, bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def _band(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(1, ncols))
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=_FONT, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=_SUB_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def _coerce(v):
    if isinstance(v, pd.Timestamp):
        return _fmt_date(v)
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
        return ""
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return round(float(v), 4)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    return v


def _write_df(ws, df: pd.DataFrame, start_row: int, *,
              green_col: str | None = None, gold_flag_col: str | None = None,
              green_flag_col: str | None = None,
              bold_total_last: bool = False, max_rows: int | None = None):
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(no data)").font = Font(name=_FONT, italic=True)
        return start_row + 1
    if max_rows is not None and len(df) > max_rows:
        df = df.head(max_rows)

    cols = list(df.columns)
    border = _thin()
    # ── header row ───────────────────────────────────────────────────────────
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.font = Font(name=_FONT, bold=True, color=_HEADER_FG, size=10)
        c.fill = PatternFill("solid", fgColor=_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    gidx  = cols.index(green_col)      if (green_col      and green_col      in cols) else None
    fidx  = cols.index(gold_flag_col)  if (gold_flag_col  and gold_flag_col  in cols) else None
    gfidx = cols.index(green_flag_col) if (green_flag_col and green_flag_col in cols) else None

    _TRUE_VALS = {"True", "true", "1", "1.0", "✓", "✓ TRADED"}

    # Pre-build coerced value matrix + flag arrays (avoids per-cell Python overhead)
    vals = df.map(_coerce).values          # shape (nrow, ncol)
    nrow, ncol = vals.shape

    # Determine per-row formatting flags
    flag_arr  = [str(vals[i, fidx])  in _TRUE_VALS for i in range(nrow)] if fidx  is not None else [False]*nrow
    green_arr = [str(vals[i, gfidx]) in _TRUE_VALS for i in range(nrow)] if gfidx is not None else [False]*nrow

    # Pre-build reusable Style objects for common cases
    _fnt_normal = Font(name=_FONT, size=10, color="000000")
    _fnt_green  = Font(name=_FONT, size=10, color=_GREEN_FG, bold=True)
    _fnt_bold   = Font(name=_FONT, size=10, color="000000", bold=True)
    _fill_green = PatternFill("solid", fgColor=_GREEN_BG)
    _fill_gold  = PatternFill("solid", fgColor=_GOLD_BG)
    _fill_alt   = PatternFill("solid", fgColor=_ALT)
    _fill_none  = PatternFill(fill_type=None)
    _align_c    = Alignment(horizontal="center", vertical="center")

    for i in range(nrow):
        r = start_row + 1 + i
        is_total  = bold_total_last and (i == nrow - 1)
        is_green  = green_arr[i]
        is_gold   = flag_arr[i] and not is_green
        is_alt    = (not is_total) and (not is_green) and (not is_gold) and (i % 2 == 1)

        row_font  = _fnt_green if is_green else (_fnt_bold if is_total else _fnt_normal)
        row_fill  = _fill_green if is_green else (_fill_gold if is_gold else
                    _fill_alt if is_alt else _fill_none)

        for j in range(ncol):
            c = ws.cell(row=r, column=j + 1, value=vals[i, j])
            c.font      = row_font
            c.border    = border
            c.alignment = _align_c
            c.fill      = row_fill

            # Green/red override for the numeric metric column
            if gidx is not None and j == gidx and not is_green:
                try:
                    fv = float(vals[i, j])
                    if fv > 0:
                        c.fill = _fill_green
                        c.font = Font(name=_FONT, size=10, color=_GREEN_FG, bold=is_total)
                    elif fv < 0:
                        c.fill = PatternFill("solid", fgColor=_RED_BG)
                        c.font = Font(name=_FONT, size=10, color=_RED_FG, bold=is_total)
                except Exception:
                    pass

    # Column widths — sample first 80 rows only
    for j, col in enumerate(cols, start=1):
        try:
            w = max([len(str(col))] + [len(str(v)) for v in vals[:80, j - 1]])
        except Exception:
            w = len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 3, 10), 36)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + 1 + nrow


# ─────────────────────────────────────────────────────────────────────────────
# Per-window leg rankings (every window, whether traded or not)
# ─────────────────────────────────────────────────────────────────────────────

def _window_leg_rankings(windows, returns_df, stock_dict, pattern, cfg,
                         bought_map, common_map, winrank_map) -> pd.DataFrame:
    metric = cfg.get("metric", "roc")
    vol_dir = cfg.get("vol_dir", "low")
    vol_type = cfg.get("vol_type", "standard")
    nn = int(cfg.get("top_n", cfg.get("topn_rise", 50)))
    _desc = cfg.get("selection_side", "top") != "bottom"   # False = ascending for bottom
    all_tk = list(stock_dict.keys())
    rows = []
    for win in windows:
        w = win["window_idx"] + 1
        for li, pid in enumerate(win["phase_ids"]):
            ttype = pattern[li] if li < len(pattern) else "Rise"
            leg = {"label": f"Leg {li+1} ({ttype})", "kind": "sched", "phase_id": pid,
                   "entry": win["entry_dates"][li], "exit": win["exit_dates"][li],
                   "trade": ttype}
            roc_map = _leg_roc_map(returns_df, leg, stock_dict, all_tk)
            vol_map = _leg_vol_map(stock_dict, all_tk, win["entry_dates"][li], win["exit_dates"][li], vol_type)
            if metric == "vol":
                # BUG FIX v30: vol sort uses vol_dir only, not inverted by selection side
                asc = (vol_dir == "low")
                order = [t for t, _ in sorted(vol_map.items(), key=lambda kv: kv[1], reverse=not asc)][:nn]
            elif metric == "volxroc":
                prod = {t: roc_map[t] * vol_map[t] for t in roc_map if t in vol_map and roc_map[t] is not None}
                order = [t for t, _ in sorted(prod.items(), key=lambda kv: kv[1], reverse=_desc)][:nn]
            elif metric == "roc_over_vol":
                rov = {t: roc_map[t] / vol_map[t] for t in roc_map
                       if t in vol_map and roc_map[t] is not None and vol_map[t] and vol_map[t] > 0}
                order = [t for t, _ in sorted(rov.items(), key=lambda kv: kv[1], reverse=_desc)][:nn]
            elif metric == "gate":
                import os
                from core.gate_system import rank_universe, load_market_features, load_quality_features, _rollup_quality, DEFAULT_PARAMS
                _gp = cfg.get("gate_params") or DEFAULT_PARAMS
                _db = cfg.get("db_path") or os.path.join("storage", "market_data.duckdb")
                _m = os.path.getmtime(_db) if os.path.exists(_db) else 0.0
                _mf = load_market_features(_db, _m)
                _qr = load_quality_features(_db, _m)
                _qroll = _rollup_quality(_qr, _gp)
                _sc = rank_universe(pd.Timestamp(win["entry_dates"][li]), all_tk, _mf, _qroll, _gp)
                _gmap = {r["ticker"]: r["combined_score"] for _, r in _sc.iterrows() if pd.notna(r["combined_score"])}
                order = [t for t, _ in sorted(_gmap.items(), key=lambda kv: kv[1], reverse=_desc)][:nn]
            else:
                order = [t for t, _ in sorted(roc_map.items(), key=lambda kv: kv[1], reverse=_desc)][:nn]
            for i, tkr in enumerate(order):
                wr = winrank_map.get((w, tkr))
                rv = roc_map.get(tkr)
                vv = vol_map.get(tkr)
                _ms = (_gmap.get(tkr) if metric == "gate"
                       else rv / vv if (metric == "roc_over_vol" and rv is not None and vv and vv > 0)
                       else rv * vv if (metric == "volxroc" and rv is not None and vv is not None)
                       else vv if metric == "vol"
                       else rv)
                rows.append({
                    "window": w,
                    "leg": f"Leg {li+1} ({ttype})",
                    "entry_date": win["entry_dates"][li],
                    "exit_date": win["exit_dates"][li],
                    "leg_rank": i + 1,
                    "ticker": tkr,
                    "roc_value": rv,
                    "vol_value": vv,
                    "metric_score": (round(float(_ms), 6) if _ms is not None else None),
                    "window_rank": wr if wr is not None else "",
                    "common": "✓" if tkr in common_map.get(w, set()) else "",
                    "bought": "✓" if tkr in bought_map.get(w, set()) else "",
                })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Day-by-day NIFTY path per cycle
# ─────────────────────────────────────────────────────────────────────────────

def _nifty_cycle_paths(cycle_df, nifty_df, fall_pct) -> pd.DataFrame:
    rows = []
    if cycle_df is None or cycle_df.empty or nifty_df is None or nifty_df.empty:
        return pd.DataFrame()
    for _, c in cycle_df.iterrows():
        b = pd.Timestamp(c["buy_trigger_date"])
        s = pd.Timestamp(c["sell_date"])
        seg = nifty_df[(nifty_df.index >= b) & (nifty_df.index <= s)]
        if seg.empty:
            continue
        buy_px = float(seg["close"].iloc[0])
        peak = -np.inf
        peak_dt = pd.Timestamp(c.get("peak_date")) if pd.notna(c.get("peak_date")) else None
        fall_dt = pd.Timestamp(c.get("fall_confirm_date")) if pd.notna(c.get("fall_confirm_date")) else None
        for dt, px in seg["close"].items():
            px = float(px)
            peak = max(peak, px)
            marker = ""
            if dt == b:
                marker = "BUY"
            elif peak_dt is not None and dt == peak_dt:
                marker = "PEAK"
            elif fall_dt is not None and dt == fall_dt:
                marker = "FALL"
            elif dt == s:
                marker = "SELL"
            rows.append({
                "cycle": int(c["cycle"]),
                "date": dt,
                "nifty_close": round(px, 2),
                "pct_from_buy": round((px - buy_px) / buy_px * 100, 2) if buy_px else 0,
                "running_peak": round(peak, 2),
                "drawdown_from_peak": round((px - peak) / peak * 100, 2) if peak > 0 else 0,
                "marker": marker,
            })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Portfolio summary helpers
# ─────────────────────────────────────────────────────────────────────────────

def _kv_block(ws, start_row, pairs, val_col=2):
    r = start_row
    for k, v in pairs:
        a = ws.cell(row=r, column=1, value=k)
        a.font = Font(name=_FONT, bold=True, size=10)
        a.alignment = Alignment(horizontal="left", indent=1)
        b = ws.cell(row=r, column=val_col, value=("" if v is None else _coerce(v)))
        b.font = Font(name=_FONT, size=10)
        b.alignment = Alignment(horizontal="left", indent=1)
        r += 1
    return r


def _inr(v):
    try:
        return round(float(v), 2)
    except Exception:
        return v


_LINK_BLUE = "0563C1"
_MAX_DETAIL_SHEETS = 200   # cap to keep the workbook a sane size


def _detail_sheet_name(c) -> str:
    return f"Cycle {int(c)} Detail"


def _hyperlink(cell, target_sheet: str, anchor: str = "A1", text=None, bold=False):
    if text is not None:
        cell.value = text
    cell.hyperlink = f"#'{target_sheet}'!{anchor}"
    cell.font = Font(name=_FONT, color=_LINK_BLUE, underline="single",
                     bold=bold, size=cell.font.size or 10)


def _build_cycle_detail(wb, c, cycle_df, per_trade_df, candidates_df,
                        eligible_ranks_df, leg_rank_df, paths, ia,
                        initial_capital, reinvest):
    """One sheet with EVERYTHING for a single cycle."""
    ws = wb.create_sheet(_detail_sheet_name(c))

    crow = cycle_df[cycle_df["cycle"] == c]
    cr = crow.iloc[0] if not crow.empty else None

    _title(ws, f"Cycle {int(c)} — Detailed Analysis", 10)
    back = ws.cell(row=2, column=1)
    _hyperlink(back, "Cycle Ledger", "A1", text="← Back to Cycle Ledger", bold=True)
    ws.row_dimensions[2].height = 18

    r = 4
    # ── Cycle Summary ──────────────────────────────────────────────────────────
    if cr is not None:
        _band(ws, r, "Cycle Summary", 10); r += 1
        pairs = [
            ("Legs used", cr.get("legs_used")),
            ("Legs window start", _fmt_date(cr.get("legs_window_start"))),
            ("Legs window end", _fmt_date(cr.get("legs_window_end"))),
            ("Buy trigger date", _fmt_date(cr.get("buy_trigger_date"))),
            ("NIFTY at buy", cr.get("nifty_at_buy")),
            ("Peak close", cr.get("peak_close")),
            ("Peak date", _fmt_date(cr.get("peak_date"))),
            ("Fall confirm date", _fmt_date(cr.get("fall_confirm_date"))),
            ("Fall close", cr.get("fall_close")),
            ("Trough close", cr.get("trough_close")),
            ("Trough date", _fmt_date(cr.get("trough_date"))),
            ("Sell date", _fmt_date(cr.get("sell_date"))),
            ("NIFTY at sell", cr.get("nifty_at_sell")),
            ("NIFTY cycle return %", cr.get("nifty_cycle_return")),
            ("Stocks held", cr.get("n_stocks")),
            ("Avg stock return %", cr.get("avg_return_pct")),
            ("Held days", cr.get("held_days")),
            ("Selection window", cr.get("legs_used") or "—"),
            ("Fall Entry window as-of date (prev peak)", _fmt_date(cr.get("fall_entry_date"))),
            ("Fall Exit  window as-of date (prev trough)", _fmt_date(cr.get("fall_exit_date"))),
            ("Vol filter basis", cr.get("vol_filter_basis") or "off"),
            ("Effective universe (selection + vol filter)", cr.get("n_effective_universe")),
            ("Common across legs", cr.get("n_common")),
            ("Selection funnel", cr.get("selection_funnel") or "—"),
            ("Status", cr.get("status")),
            ("Reason", cr.get("reason") or "—"),
        ]
        r = _kv_block(ws, r, pairs) + 1
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 24

    # ── Investment snapshot for this cycle ─────────────────────────────────────
    inv_row = None
    if ia is not None and not ia["window_table"].empty:
        wt = ia["window_table"]
        m = wt[wt["window"] == c]
        if not m.empty:
            inv_row = m.iloc[0]
    _band(ws, r, "Investment Analysis — this cycle", 10); r += 1
    if inv_row is not None:
        cap = float(inv_row.get("per_window_capital", 0) or 0)
        nstk = int(inv_row.get("n_stocks", 0) or 0)
        pairs = [
            ("Capital deployed (₹)", _inr(cap)),
            ("Per-stock allocation (₹)", _inr(cap / nstk) if nstk else "—"),
            ("Cycle return %", inv_row.get("window_return_pct")),
            ("Cycle profit (₹)", _inr(inv_row.get("window_profit_inr"))),
            ("Equity after cycle (₹)", _inr(inv_row.get("equity_inr"))),
            ("Cumulative profit (₹)", _inr(inv_row.get("cumulative_profit_inr"))),
            ("Avg alpha %", inv_row.get("avg_alpha")),
            ("Avg holding days", inv_row.get("avg_days_held")),
            ("Reinvest profits", "Yes" if reinvest else "No"),
        ]
        r = _kv_block(ws, r, pairs) + 1
    else:
        ws.cell(row=r, column=1, value="(needs ≥ 2 completed cycles for portfolio equity)") \
            .font = Font(name=_FONT, italic=True)
        r += 2

    # ── Traded stocks (with per-stock allocation & profit) ─────────────────────
    _band(ws, r, "Stocks Bought & Sold (P&L)", 10); r += 1
    if per_trade_df is not None and not per_trade_df.empty:
        tt = per_trade_df[per_trade_df["cycle"] == c].copy()
        if not tt.empty:
            alloc = None
            if inv_row is not None and int(inv_row.get("n_stocks", 0) or 0) > 0:
                alloc = float(inv_row["per_window_capital"]) / int(inv_row["n_stocks"])
            for dc in ["entry_date", "exit_date"]:
                tt[dc] = tt[dc].map(_fmt_date)
            if alloc is not None:
                tt["alloc_inr"] = round(alloc, 2)
                tt["profit_inr"] = (alloc * tt["return_pct"] / 100.0).round(2)
            cols = [x for x in ["ticker", "entry_date", "entry_price",
                    "exit_date", "exit_price", "return_pct", "nifty_return", "alpha", "days_held",
                    "alloc_inr", "profit_inr"] if x in tt.columns]
            disp = tt[cols].rename(columns=lambda x: x.replace("_", " ").title())
            r = _write_df(ws, disp, r, green_col="Return Pct") + 1
        else:
            ws.cell(row=r, column=1, value="(no trades)").font = Font(name=_FONT, italic=True); r += 2
    else:
        ws.cell(row=r, column=1, value="(no trades)").font = Font(name=_FONT, italic=True); r += 2

    # ── Cycle Candidates (common ranking) ──────────────────────────────────────
    _band(ws, r, "Cycle Candidates — common ranking (leg rank + window rank)", 10); r += 1
    if candidates_df is not None and not candidates_df.empty and "cycle" in candidates_df.columns:
        cd = candidates_df[candidates_df["cycle"] == c].copy()
        if not cd.empty:
            cd = cd.drop(columns=[x for x in ["window"] if x in cd.columns])
            cd = cd.rename(columns=lambda x: x.replace("_", " ").title())
            r = _write_df(ws, cd, r, gold_flag_col="Selected To Buy") + 1
        else:
            ws.cell(row=r, column=1, value="(none)").font = Font(name=_FONT, italic=True); r += 2
    else:
        ws.cell(row=r, column=1, value="(none)").font = Font(name=_FONT, italic=True); r += 2

    # ── Cycle Leg Rankings ─────────────────────────────────────────────────────
    _vt_lbl = "Downside" if True else "Std"  # placeholder - actual in sheet title
    _side_n3 = "Bot-N" if cfg.get("selection_side", "top") == "bottom" else "Top-N"
    _band(ws, r, f"Leg Rankings ({_side_n3} per leg, with Ranking Score)", 10); r += 1
    if leg_rank_df is not None and not leg_rank_df.empty and "cycle" in leg_rank_df.columns:
        lr = leg_rank_df[leg_rank_df["cycle"] == c].copy()
        if not lr.empty:
            for dc in ["entry_date", "exit_date"]:
                if dc in lr.columns:
                    lr[dc] = lr[dc].map(_fmt_date)
            keep = [x for x in ["leg", "entry_date", "exit_date", "leg_rank", "ticker",
                    "roc_value", "vol_value", "metric_score",
                    "beta_nifty", "corr_nifty", "beta_rank",
                    "window_rank", "common", "bought"] if x in lr.columns]
            lr = lr[keep].rename(columns=lambda x: x.replace("_", " ").title())
            r = _write_df(ws, lr, r, green_flag_col="Bought") + 1
        else:
            ws.cell(row=r, column=1, value="(none)").font = Font(name=_FONT, italic=True); r += 2
    else:
        ws.cell(row=r, column=1, value="(none)").font = Font(name=_FONT, italic=True); r += 2

    # ── Eligible Stock Ranking (quartiles, traded highlighted) ─────────────────
    _band(ws, r, "Eligible Stock Ranking — NIFTY 500 universe · Q4..Q1 · green = traded", 10); r += 1
    if eligible_ranks_df is not None and not eligible_ranks_df.empty:
        ev = eligible_ranks_df[eligible_ranks_df["cycle"] == c].copy()
        if not ev.empty:
            for dc in ["entry_date", "exit_date"]:
                if dc in ev.columns:
                    ev[dc] = ev[dc].map(_fmt_date)
            ev["Traded?"] = ev["traded"].map(lambda v: "✓ TRADED" if bool(v) else "—")
            cols = [x for x in ["rank", "ticker", "quartile",
                    "entry_date", "exit_date",
                    "entry_price", "exit_price", "return_pct", "nifty_return", "alpha",
                    "n_eligible", "Traded?"] if x in ev.columns]
            ev = ev[cols].sort_values("rank").rename(
                columns=lambda x: x.replace("_", " ").title() if x != "Traded?" else x)
            r = _write_df(ws, ev, r, green_flag_col="Traded?", max_rows=4000) + 1
        else:
            ws.cell(row=r, column=1, value="(none)").font = Font(name=_FONT, italic=True); r += 2
    else:
        ws.cell(row=r, column=1, value="(none)").font = Font(name=_FONT, italic=True); r += 2

    # ── NIFTY path for this cycle ──────────────────────────────────────────────
    _band(ws, r, "NIFTY Path (day-by-day) — running peak, drawdown, markers", 10); r += 1
    if paths is not None and not paths.empty:
        pp = paths[paths["cycle"] == c].copy()
        if not pp.empty:
            pp["date"] = pp["date"].map(_fmt_date)
            pp = pp.drop(columns=[x for x in ["cycle"] if x in pp.columns])
            pp = pp.rename(columns=lambda x: x.replace("_", " ").title())
            r = _write_df(ws, pp, r, green_col="Pct From Buy", max_rows=4000) + 1
        else:
            ws.cell(row=r, column=1, value="(no path)").font = Font(name=_FONT, italic=True); r += 1
    else:
        ws.cell(row=r, column=1, value="(no path)").font = Font(name=_FONT, italic=True); r += 1

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Quartile-aware year-wise / period summary (mirrors the short-term system)
# ─────────────────────────────────────────────────────────────────────────────

_SUMMARY_HEADERS = ["Period / Year", "Trades", "Q1 %", "Q2 %", "Q3 %", "Q4 %",
                    "Q3+Q4 #", "Q3+Q4 %", "Gross %", "Avg %", "Win %",
                    "Win Avg %", "Loss Avg %", "Avg Alpha %", "Avg NIFTY %",
                    "Best %", "Worst %"]


def _attach_qnum(per_trade_df: pd.DataFrame, eligible_ranks_df: pd.DataFrame) -> pd.DataFrame:
    """Attach `_qnum` (1..4, 4 = top) to each trade from the eligible-stock
    (buy→sell return) quartile, keyed by (cycle, ticker)."""
    df = per_trade_df.copy()
    qmap = {}
    if eligible_ranks_df is not None and not eligible_ranks_df.empty \
            and "quartile" in eligible_ranks_df.columns:
        for _, r in eligible_ranks_df.iterrows():
            qmap[(int(r["cycle"]), str(r["ticker"]))] = str(r.get("quartile", ""))

    def _qnum(row):
        lbl = qmap.get((int(row.get("cycle", -1)), str(row.get("ticker", ""))), "")
        head = lbl.split()[0] if lbl.split() else ""
        return int(head[1]) if head[:1] == "Q" and head[1:2].isdigit() else 0

    df["_qnum"] = df.apply(_qnum, axis=1)
    df["_yr"] = pd.to_datetime(df["entry_date"], errors="coerce").dt.year
    return df


def _trade_metrics_q(df: pd.DataFrame) -> dict:
    base = {k: 0 for k in ["n", "q1", "q2", "q3", "q4", "q34"]}
    base.update({k: 0.0 for k in ["gross", "avg", "winpct", "winavg", "lossavg",
                                   "alpha", "nifty", "best", "worst",
                                   "q1p", "q2p", "q3p", "q4p", "q34p"]})
    if df is None or df.empty:
        return base
    ret = pd.to_numeric(df["return_pct"], errors="coerce").dropna()
    if ret.empty:
        return base
    wins, losses = ret[ret > 0], ret[ret < 0]
    _m = lambda s: round(float(s.mean()), 2) if len(s) else 0.0
    alpha = pd.to_numeric(df.get("alpha", pd.Series(dtype=float)), errors="coerce").dropna()
    nifty = pd.to_numeric(df.get("nifty_return", pd.Series(dtype=float)), errors="coerce").dropna()
    m = {"n": int(len(ret)), "gross": round(float(ret.sum()), 2), "avg": _m(ret),
         "winpct": round(float((ret > 0).mean() * 100), 1), "winavg": _m(wins),
         "lossavg": _m(losses), "alpha": _m(alpha), "nifty": _m(nifty),
         "best": round(float(ret.max()), 2), "worst": round(float(ret.min()), 2)}
    n = len(df)
    qn = pd.to_numeric(df.get("_qnum", pd.Series([0]*n)), errors="coerce").fillna(0).astype(int)
    c1, c2, c3, c4 = (int((qn == k).sum()) for k in (1, 2, 3, 4))
    c34 = c3 + c4
    pct = lambda c: round(c / n * 100, 1) if n else 0.0
    m.update({"q1": c1, "q2": c2, "q3": c3, "q4": c4, "q34": c34,
              "q1p": pct(c1), "q2p": pct(c2), "q3p": pct(c3), "q4p": pct(c4), "q34p": pct(c34)})
    return m


def _summary_rows_q(dfq: pd.DataFrame, kind: str) -> list[tuple[str, dict]]:
    out = []
    if dfq is None or dfq.empty or "_yr" not in dfq.columns:
        return out
    yrs = dfq["_yr"].dropna()
    if yrs.empty:
        return out
    end_yr = int(yrs.max())
    if kind == "period":
        start_yr = int(yrs.min())
        mids = sorted({start_yr, max(start_yr, end_yr - 8), max(start_yr, end_yr - 3)})
        for s in mids:
            sub = dfq[(dfq["_yr"] >= s) & (dfq["_yr"] <= end_yr)]
            out.append((f"{s}–{end_yr}", _trade_metrics_q(sub)))
    else:
        for yr in sorted(yrs.unique()):
            out.append((str(int(yr)), _trade_metrics_q(dfq[dfq["_yr"] == yr])))
        out.append(("All Years", _trade_metrics_q(dfq)))
    return out


def _write_summary_q(ws, start_row: int, title: str, rows: list[tuple[str, dict]]) -> int:
    ncols = len(_SUMMARY_HEADERS)
    _band(ws, start_row, title, ncols)
    r = start_row + 1
    border = _thin()
    for c, h in enumerate(_SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(name=_FONT, bold=True, color=_HEADER_FG, size=9)
        cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = 11 if c > 1 else 14
    r += 1
    if not rows:
        ws.cell(row=r, column=1, value="No trades available.").font = Font(name=_FONT, italic=True)
        return r + 2
    for label, m in rows:
        is_total = label == "All Years"
        vals = [label, m["n"], m["q1p"], m["q2p"], m["q3p"], m["q4p"], m["q34"], m["q34p"],
                m["gross"], m["avg"], m["winpct"], m["winavg"], m["lossavg"],
                m["alpha"], m["nifty"], m["best"], m["worst"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=_coerce(v))
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name=_FONT, size=9, bold=is_total or c == 1)
            if c == 1:
                cell.fill = PatternFill("solid", fgColor="DDEBF7")
            elif 3 <= c <= 6:
                cell.fill = PatternFill("solid", fgColor=_GOLD_BG)
            elif c in (7, 8):
                cell.fill = PatternFill("solid", fgColor="DDEBF7")
            elif c in (9, 10) and isinstance(v, (int, float)):
                cell.fill = PatternFill("solid", fgColor=_GREEN_BG if v >= 0 else _RED_BG)
                cell.font = Font(name=_FONT, size=9, color=_GREEN_FG if v >= 0 else _RED_FG,
                                 bold=is_total)
        r += 1
    return r + 1


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def generate_momentum_excel(
    per_trade_df: pd.DataFrame,
    cycle_df: pd.DataFrame,
    candidates_df: pd.DataFrame,
    status_df: pd.DataFrame,
    phases: pd.DataFrame,
    config: dict,
    *,
    eligible_ranks_df: pd.DataFrame = None,
    leg_rank_df: pd.DataFrame = None,
    nifty_df: pd.DataFrame = None,
    stock_dict: dict = None,
    returns_df: pd.DataFrame = None,
    initial_capital: float = 100_000.0,
    reinvest: bool = True,
) -> bytes:
    cfg = config or {}
    pattern = cfg.get("pattern", [])
    wb = Workbook()

    # ── investment analysis (shared) ───────────────────────────────────────────
    ia = None
    if per_trade_df is not None and not per_trade_df.empty \
            and per_trade_df["window_idx"].nunique() >= 2:
        ia = compute_investment_analysis(per_trade_df, initial_capital=initial_capital,
                                         alloc_mode="equal", reinvest=reinvest)

    # bought / common / window-rank maps keyed by source window number
    bought_map: dict[int, set] = {}
    common_map: dict[int, set] = {}
    winrank_map: dict[tuple, int] = {}
    if per_trade_df is not None and not per_trade_df.empty and "source_window" in per_trade_df.columns:
        for w, g in per_trade_df.groupby("source_window"):
            bought_map[int(w)] = set(g["ticker"])
    if candidates_df is not None and not candidates_df.empty and "window" in candidates_df.columns:
        for w, g in candidates_df.groupby("window"):
            common_map[int(w)] = set(g["ticker"])
        if "common_rank" in candidates_df.columns:
            for _, r in candidates_df.iterrows():
                winrank_map[(int(r["window"]), str(r["ticker"]))] = int(r["common_rank"])

    # ══ ① Portfolio Summary ════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Portfolio Summary"  # always fixed name for Excel nav
    _m = cfg.get("metric", "roc")
    _vt = cfg.get("vol_type", "standard")
    _use_fixed = cfg.get("use_fixed_fall", False)
    _top_n = cfg.get("top_n", cfg.get("vol_filter_n", 50))
    _m_lbl = (
        "Off (Vol-filter only)" if _m == "off"
        else "ROC" if _m == "roc"
        else f"Vol-{'DS' if _vt=='downside' else 'Std'} ({cfg.get('vol_dir','low')})" if _m == "vol"
        else f"ROC×Vol-{'DS' if _vt=='downside' else 'Std'}" if _m == "volxroc"
        else f"ROC/Vol-{'DS' if _vt=='downside' else 'Std'}" if _m == "roc_over_vol"
        else f"Beta/Vol-{'DS' if _vt=='downside' else 'Std'}" if _m == "beta_over_vol"
        else f"Beta×Vol-{'DS' if _vt=='downside' else 'Std'}" if _m == "beta_x_vol"
        else "Std Dev / Downside Vol (σ÷DV)" if _m == "sd_over_dv"
        else f"Both-ROC∩Vol-{'DS' if _vt=='downside' else 'Std'}({cfg.get('vol_dir','low')})"
    )
    _k   = cfg.get("top_k", "?")
    _np  = cfg.get("nifty_pct", "?")
    _fp  = cfg.get("fall", "?")
    _hd  = cfg.get("max_hold", "?")
    _run_name = _make_run_name(cfg)
    _ps_title = f"S³ Portfolio Report — {_run_name}"
    _title(ws, _ps_title, 4)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 26

    _band(ws, 3, "Strategy Configuration", 4)
    _metric = cfg.get("metric", "roc")
    _vt = cfg.get("vol_type", "standard")
    _use_fixed = cfg.get("use_fixed_fall", False)
    _top_n = cfg.get("top_n", cfg.get("vol_filter_n", 50))
    _fe_days = cfg.get("fall_entry_days", 100)
    _fx_days = cfg.get("fall_exit_days", 100)
    _vf = cfg.get("vol_filter", "off")
    _vf_wins = cfg.get("vf_windows") or [("buy", int(cfg.get("vol_filter_lookback_days", 100)))]
    _off_days = _vf_wins[0][1] if _vf_wins else 100
    _is_off = (_metric == "off")
    _vol_type_txt = ("Downside: sqrt(mean(min(log-ret,0)²))×sqrt(252)"
                     if _vt == "downside" else
                     "Standard: std(ln(Pₜ/Pₜ₋₁))×sqrt(252)")
    _metric_txt = (
        "Off (volatility filter only)" if _metric == "off"
        else "ROC" if _metric == "roc"
        else f"Volatility ({cfg.get('vol_dir', 'low')}, " + ("Bot-N" if cfg.get("selection_side","top")=="bottom" else "Top-N") + ")" if _metric == "vol"
        else "Vol × ROC (product)" if _metric == "volxroc"
        else "ROC / Volatility (Sharpe-style)" if _metric == "roc_over_vol"
        else "Beta / Volatility (β÷vol)" if _metric == "beta_over_vol"
        else "Beta × Volatility (β×vol)" if _metric == "beta_x_vol"
        else "Std Dev / Downside Vol (σ÷DV ratio)" if _metric == "sd_over_dv"
        else f"Both — ROC ∩ Volatility ({cfg.get('vol_dir', 'low')}, " + ("Bot-N" if cfg.get("selection_side","top")=="bottom" else "Top-N") + ")")
    _sel_method = (
        f"Pure {_off_days}-day vol — " + ('bot' if cfg.get('selection_side','top')=='bottom' else 'top') + f"-{cfg.get('vol_filter_n', _top_n)} "
        f"{'calmest' if _vf == 'low' else 'most-volatile'} per window → intersection → buy K"
        if _is_off else
        f"Fall Entry: {_fe_days}d window ending at prev cycle PEAK; "
        f"Fall Exit: {_fx_days}d window ending at prev cycle TROUGH. "
        f"All stocks ranked and bought without arbitrary Aux filtering."
        if _use_fixed else
        "Auto: last two scheduled legs (Rise + dynamic fall window). "
        "All stocks ranked and bought without arbitrary Aux filtering.")
    _fixed_days_row = (
        f"Fall Entry = {_fe_days}d @ prev peak  |  Fall Exit = {_fx_days}d @ prev trough"
        if _use_fixed else "—")
    cfg_pairs = [
        ("Run name (settings)", _run_name),
        ("Selection window", _sel_method),
        ("Ranking metric", _metric_txt),
        ("Volatility type", _vol_type_txt),
        ("Selection side", "Top-N (best) & common" if cfg.get("selection_side", "top") == "top"
         else "Bottom-N (worst) & common"),
        (("Top-N" if cfg.get("selection_side","top") == "top" else "Bottom-N") + " per window", str(_top_n)),
        ("Fixed window lookback", _fixed_days_row if _use_fixed else "N/A"),
        ("Common stocks bought (K)", cfg.get("top_k")),
        ("NIFTY % (BUY rise & SELL recovery)", cfg.get("nifty_pct")),
        ("Fall % from running peak", cfg.get("fall")),
        ("Keep-holding window (yrs)", cfg.get("max_hold")),
        ("Exact Trigger Mode", "Yes — Sell at n% fall; Buy at n% recovery (separate dates)"
         if cfg.get("exact_trigger_mode") else "No — Sell & next Buy on same date (recovery)"),
        ("Cycle model", "Exact Trigger (sell≠buy date)" if cfg.get("exact_trigger_mode")
         else "Continuous (sell date = next buy date)"),
        ("Initial capital (₹)", _inr(initial_capital)),
        ("Reinvest profits", "Yes" if reinvest else "No"),
    ]
    end_r = _kv_block(ws, 4, cfg_pairs)

    _band(ws, end_r + 1, "Headline Portfolio Statistics", 4)
    r0 = end_r + 2
    n_cyc = int(len(cycle_df)) if cycle_df is not None else 0
    n_done = int((cycle_df["status"] == "Completed cycle").sum()) if (cycle_df is not None and not cycle_df.empty) else 0
    n_trades = int(len(per_trade_df)) if per_trade_df is not None else 0
    stat_pairs = [
        ("Total cycles", n_cyc),
        ("Completed cycles", n_done),
        ("Total trades", n_trades),
    ]
    if ia is not None:
        m = ia["metrics"]
        stat_pairs += [
            ("Final equity (₹)", _inr(m["final_equity"])),
            ("Total P/L (₹)", _inr(m["total_pl"])),
            ("Total return %", round(m["total_pl_pct"], 2)),
            ("CAGR %", None if np.isnan(m["cagr"]) else round(m["cagr"], 2)),
            ("Max drawdown %", round(m["mdd_pct"], 2)),
            ("CAR / MDD (Calmar)", None if np.isnan(m["calmar"]) else round(m["calmar"], 2)),
            ("Sharpe", None if np.isnan(m["sharpe"]) else round(m["sharpe"], 2)),
            ("Win rate %", round(m["win_rate"], 2)),
            ("Alpha win rate %", None if np.isnan(m["alpha_win_rate"]) else round(m["alpha_win_rate"], 2)),
            ("Avg profit / trade (₹)", _inr(m["avg_profit_trade"])),
            ("Avg holding days", None if np.isnan(m["avg_holding_days"]) else round(m["avg_holding_days"], 1)),
            ("Years covered", None if np.isnan(m.get("years", np.nan)) else round(m["years"], 2)),
        ]
        bw, wwn = m.get("best_window"), m.get("worst_window")
        if bw:
            stat_pairs.append(("Best cycle profit (₹)", _inr(bw["profit"])))
        if wwn:
            stat_pairs.append(("Worst cycle profit (₹)", _inr(wwn["profit"])))
    else:
        stat_pairs.append(("Investment analysis", "Needs ≥ 2 completed cycles"))
    _kv_block(ws, r0, stat_pairs)

    # ══ ② Cycle Ledger ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Cycle Ledger")
    detail_cycles = []
    if cycle_df is not None and not cycle_df.empty:
        cyc = cycle_df.copy()
        for dc in ["legs_window_start", "legs_window_end", "buy_trigger_date", "peak_date",
                   "fall_confirm_date", "sell_date", "fall_entry_date", "fall_exit_date"]:
            if dc in cyc.columns:
                cyc[dc] = cyc[dc].map(_fmt_date)
        cyc = cyc.rename(columns=lambda c: c.replace("_", " ").title())
        cyc.insert(0, "Open", ["▶ Detail"] * len(cyc))   # link column
        _title(ws2, "Cycle Ledger — click ‘▶ Detail’ (or the cycle number) to open that "
               "cycle's full analysis", len(cyc.columns))
        _write_df(ws2, cyc, 3, green_col="Avg Return Pct")

        # hyperlinks: column 1 (Open) and column 2 (Cycle) → per-cycle detail sheet
        order = list(cycle_df["cycle"])
        detail_cycles = [int(x) for x in order[:_MAX_DETAIL_SHEETS]]
        detail_set = set(detail_cycles)
        for i, cval in enumerate(order):
            cval = int(cval)
            if cval not in detail_set:
                continue
            data_row = 4 + i
            tgt = _detail_sheet_name(cval)
            _hyperlink(ws2.cell(row=data_row, column=1), tgt, "A1", text="▶ Detail", bold=True)
            _hyperlink(ws2.cell(row=data_row, column=2), tgt, "A1", text=cval)
    else:
        _title(ws2, "Cycle Ledger", 2)
        ws2.cell(row=3, column=1, value="(no cycles)")

    # ══ ③ Cycle Status ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Cycle Status")
    if status_df is not None and not status_df.empty:
        sd = status_df.copy()
        for dc in ["pattern_start", "pattern_end", "buy_trigger_date", "sell_date"]:
            if dc in sd.columns:
                sd[dc] = sd[dc].map(_fmt_date)
        sd = sd.rename(columns=lambda c: c.replace("_", " ").title())
        _title(ws3, "Cycle Status — every pattern window (traded or skipped)", len(sd.columns))
        _write_df(ws3, sd, 3)
    else:
        _title(ws3, "Cycle Status", 2)
        ws3.cell(row=3, column=1, value="(no windows)")

    # ══ ④ Cycle Candidates ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Cycle Candidates")
    if candidates_df is not None and not candidates_df.empty:
        cd = candidates_df.copy().rename(columns=lambda c: c.replace("_", " ").title())
        cd = cd.rename(columns={
            "Mean Sd Over Dv":  "Mean σ÷DV Ratio",
            "Mean Sd Std Vol":  "Mean Std Dev Vol",
            "Mean Sd Dv Vol":   "Mean Downside Vol",
        })
        _title(ws4, "Cycle Candidates — pure-ranking common stocks per cycle", len(cd.columns))
        _write_df(ws4, cd, 3, gold_flag_col="Selected To Buy")
    else:
        _title(ws4, "Cycle Candidates", 2)
        ws4.cell(row=3, column=1, value="(no candidates)")

    # ══ ④b Eligible Stock Ranking (verification) ═════════════════════════════════
    wsE = wb.create_sheet("Eligible Stock Ranking")
    if eligible_ranks_df is not None and not eligible_ranks_df.empty:
        er = eligible_ranks_df.copy()
        for dc in ["entry_date", "exit_date"]:
            if dc in er.columns:
                er[dc] = er[dc].map(_fmt_date)
        order = [c for c in ["cycle", "source_window", "rank", "ticker", "quartile",
                 "entry_date", "exit_date", "entry_price", "exit_price",
                 "return_pct", "nifty_return", "alpha", "n_eligible", "Traded?"]
                 if c in er.columns]
        er = er[order].sort_values(["cycle", "rank"])
        er = er.rename(columns=lambda c: c.replace("_", " ").title() if c != "Traded?" else c)
        _title(wsE, "Eligible Stock Ranking — ALL eligible stocks ranked by actual Buy→Sell "
               "return  |  Quartile Q4..Q1 (Q4 = best 25%)  |  green = actually bought/sold",
               len(er.columns))
        _write_df(wsE, er, 3, green_flag_col="Traded?", max_rows=30000)
    else:
        _title(wsE, "Eligible Stock Ranking", 2)
        wsE.cell(row=3, column=1, value="(no executed cycles)")

    # ══ ⑤ Cycle Leg Rankings (the rolling legs used per cycle) ═══════════════════
    ws5 = wb.create_sheet("Cycle Leg Rankings")
    leg_rk = leg_rank_df.copy() if (leg_rank_df is not None and not leg_rank_df.empty) else pd.DataFrame()
    if leg_rk.empty and cfg.get("metric") != "off" and (returns_df is not None and stock_dict is not None
                         and phases is not None and pattern):
        # fallback: pattern-window leg rankings
        try:
            windows = find_pattern_windows(phases, pattern)
            leg_rk = _window_leg_rankings(windows, returns_df, stock_dict, pattern, cfg,
                                          bought_map, common_map, winrank_map)
        except Exception:
            leg_rk = pd.DataFrame()
    if not leg_rk.empty:
        # ── Normalize leg dates (safety net: engine now filters at source) ──────
        # For each (cycle/window, leg) group keep only rows whose entry_date AND
        # exit_date match the group's mode (most common) value.
        _grp_cols = [c for c in ["cycle", "window", "leg"] if c in leg_rk.columns]
        if _grp_cols and "entry_date" in leg_rk.columns and "exit_date" in leg_rk.columns:
            _norm_parts = []
            for _, _g in leg_rk.groupby(_grp_cols, sort=False):
                _em = _g["entry_date"].mode()
                _xm = _g["exit_date"].mode()
                if _em.empty or _xm.empty:
                    continue
                _em, _xm = _em.iloc[0], _xm.iloc[0]
                # reset_index prevents "identically-labeled Series" comparison error
                _gr = _g.reset_index(drop=True)
                _keep = _gr[(_gr["entry_date"] == _em) & (_gr["exit_date"] == _xm)].copy()
                _norm_parts.append(_keep)
            leg_rk = pd.concat(_norm_parts).reset_index(drop=True) if _norm_parts else leg_rk
        # ─────────────────────────────────────────────────────────────────────────
        disp = leg_rk.copy()
        for dc in ["entry_date", "exit_date"]:
            if dc in disp.columns:
                disp[dc] = disp[dc].map(_fmt_date)
        _exp_vt = cfg.get("vol_type", "standard")
        _exp_metric = cfg.get("metric", "roc")
        _exp_vol_lbl = "Downside Vol" if _exp_vt == "downside" else "Vol Value"
        _exp_score_lbl = {
            "roc_over_vol": "ROC/Vol Score",
            "volxroc":      "ROC×Vol Score",
            "vol":          "Vol Score",
            "roc":          "ROC Score",
            "both":         "ROC Score",
            "beta":         "Beta (NIFTY)",
            "beta_over_vol": "Beta/Vol Score",
            "beta_x_vol":   "Beta×Vol Score",
            "sd_over_dv":   "Std/Downside Ratio",
        }.get(_exp_metric, "Score")
        keep = [c for c in ["cycle", "window", "leg", "entry_date", "exit_date",
                "leg_rank", "ticker", "roc_value", "vol_value",
                "sd_std_vol", "sd_dv_vol", "metric_score",
                "beta_nifty", "corr_nifty", "beta_rank",
                "window_rank", "common", "bought"]
                if c in disp.columns]
        disp = disp[keep].rename(columns=lambda c: c.replace("_", " ").title())
        # Override specific column names for clarity
        disp = disp.rename(columns={
            "Vol Value":    _exp_vol_lbl,
            "Metric Score": _exp_score_lbl,
            "Sd Std Vol":   "Std Volatility",
            "Sd Dv Vol":    "Downside Volatility",
        })
        if cfg.get("metric") == "off":
            _sn2 = "Bot-N" if cfg.get("selection_side", "top") == "bottom" else "Top-N"
            _wdays = cfg.get("vf_windows", [("buy", 100)])[0][1] if cfg.get("vf_windows") else 100
            _title(ws5, f"Volatility Window Rankings — {_sn2} per {_wdays}-day window by ln "
                   + f"volatility  ·  exact window dates & trading days  ·  ✓ = common (in {_sn2} of every window) / bought", len(disp.columns))
        else:
            _mlabel = ("ROC" if cfg.get("metric") == "roc"
                       else "Volatility" if cfg.get("metric") == "vol"
                       else "Vol × ROC" if cfg.get("metric") == "volxroc"
                       else "ROC / Volatility" if cfg.get("metric") == "roc_over_vol"
                       else "Beta (NIFTY ln-ret, base-day prepended) — lowest β selected" if cfg.get("metric") == "beta"
                       else "Beta / Volatility (β÷vol)" if cfg.get("metric") == "beta_over_vol"
                       else "Beta × Volatility (β×vol)" if cfg.get("metric") == "beta_x_vol"
                       else "Std Dev / Downside Vol (σ÷DV)" if cfg.get("metric") == "sd_over_dv"
                       else "ROC + Volatility")
            _fixed_leg_sfx = (f"  ·  Fixed windows: FallEntry={cfg.get('fall_entry_days',100)}d@prev-peak, "
                               f"FallExit={cfg.get('fall_exit_days',100)}d@prev-trough"
                               if cfg.get("use_fixed_fall") else "")
            _side_n = "Bot-N" if cfg.get("selection_side", "top") == "bottom" else "Top-N"
            _title(ws5, f"Cycle Leg Rankings — {_side_n} per leg by " + _mlabel
                   + "  ·  ROC % and leg volatility shown for every stock  ·  ✓ = common / bought "
                   "(quartiles are on 'Eligible Stock Ranking')" + _fixed_leg_sfx, len(disp.columns))
        _write_df(ws5, disp, 3, green_flag_col="Bought")
    else:
        _title(ws5, "Cycle Leg Rankings", 2)
        ws5.cell(row=3, column=1, value=(
            "Ranking is OFF — selection is pure volatility filter only."
            if cfg.get("metric") == "off" else "(no cycles executed)"))

    # ══ ⑥ Per-Cycle Equity ═══════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Per-Cycle Equity")
    if ia is not None and not ia["window_table"].empty:
        wt = ia["window_table"].copy()
        for dc in ["entry_date", "exit_date"]:
            if dc in wt.columns:
                wt[dc] = wt[dc].map(_fmt_date)
        keep = [c for c in ["window", "entry_date", "exit_date", "n_stocks",
                "per_window_capital", "window_return_pct", "window_profit_inr",
                "cumulative_profit_inr", "equity_inr", "avg_alpha", "avg_days_held"]
                if c in wt.columns]
        wt = wt[keep].rename(columns={"window": "cycle"})
        wt = wt.rename(columns=lambda c: c.replace("_", " ").title())
        _title(ws6, "Per-Cycle Equity — capital, return, profit, running equity", len(wt.columns))
        _write_df(ws6, wt, 3, green_col="Window Profit Inr")
    else:
        _title(ws6, "Per-Cycle Equity", 2)
        ws6.cell(row=3, column=1, value="(needs ≥ 2 completed cycles)")

    # ══ ⑦ Common Stocks P&L ══════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Common Stocks P&L")
    if per_trade_df is not None and not per_trade_df.empty:
        p = per_trade_df.copy()
        for dc in ["entry_date", "exit_date"]:
            p[dc] = p[dc].map(_fmt_date)
        p["P&L"] = p["return_pct"].map(lambda v: "PROFIT" if (pd.notna(v) and v > 0) else "LOSS")
        order = [c for c in ["cycle", "source_window", "ticker", "entry_date", "entry_price",
                 "exit_date", "exit_price", "return_pct", "nifty_return", "alpha",
                 "days_held", "status", "P&L"] if c in p.columns]
        p = p[order].sort_values("alpha", ascending=False) if "alpha" in p.columns else p[order]
        p = p.rename(columns=lambda c: c.replace("_", " ").title() if c != "P&L" else c)
        _title(ws7, "Common Stocks — Entry / Exit / P&L per cycle", len(p.columns))
        _write_df(ws7, p, 3, green_col="Return Pct")
    else:
        _title(ws7, "Common Stocks P&L", 2)
        ws7.cell(row=3, column=1, value="(no trades)")

    # ══ ⑧ Trade Log ══════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Trade Log")
    if per_trade_df is not None and not per_trade_df.empty:
        t = per_trade_df.copy()
        for dc in ["entry_date", "exit_date", "buy_trigger_date", "sell_date_cycle",
                   "peak_date", "fall_confirm_date", "pattern_start", "pattern_end",
                   "buy_phase_start"]:
            if dc in t.columns:
                t[dc] = t[dc].map(_fmt_date)
        order = [c for c in ["cycle", "source_window", "ticker", "buy_trigger_date",
                 "nifty_buy_base", "nifty_at_buy", "peak_close", "peak_date",
                 "fall_confirm_date", "trough_close", "sell_date_cycle", "nifty_at_sell",
                 "entry_date", "entry_price", "exit_date", "exit_price", "return_pct",
                 "days_held", "status"] if c in t.columns]
        t = t[order].rename(columns=lambda c: c.replace("_", " ").title())
        _title(ws8, "Trade Log — every executed momentum trade", len(t.columns))
        _write_df(ws8, t, 3, green_col="Return Pct")
    else:
        _title(ws8, "Trade Log", 2)
        ws8.cell(row=3, column=1, value="(no trades)")

    # ══ ⑨ Stock Summary ══════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("Stock Summary")
    if per_trade_df is not None and not per_trade_df.empty:
        rows = []
        for tkr, g in per_trade_df.groupby("ticker"):
            ret = g["return_pct"].dropna()
            alp = g["alpha"].dropna() if "alpha" in g.columns else pd.Series(dtype=float)
            rows.append({
                "ticker": tkr,
                "n_trades": int(len(g)),
                "n_cycles": int(g["cycle"].nunique()) if "cycle" in g.columns else int(len(g)),
                "avg_return": round(float(ret.mean()), 2) if len(ret) else None,
                "avg_alpha": round(float(alp.mean()), 2) if len(alp) else None,
                "win_rate": round(float((ret > 0).mean() * 100), 1) if len(ret) else None,
                "best_return": round(float(ret.max()), 2) if len(ret) else None,
                "worst_return": round(float(ret.min()), 2) if len(ret) else None,
                "avg_days_held": round(float(g["days_held"].mean()), 1) if "days_held" in g.columns else None,
            })
        ss = (pd.DataFrame(rows).sort_values("avg_alpha", ascending=False, na_position="last")
              if rows else pd.DataFrame())
        ss = ss.rename(columns=lambda c: c.replace("_", " ").title())
        _title(ws9, "Stock Summary — per-ticker aggregated statistics", len(ss.columns) if not ss.empty else 2)
        _write_df(ws9, ss, 3, green_col="Avg Alpha")
    else:
        _title(ws9, "Stock Summary", 2)
        ws9.cell(row=3, column=1, value="(no trades)")

    # ══ ⑩ NIFTY Cycle Path ════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("NIFTY Cycle Path")
    paths = _nifty_cycle_paths(cycle_df, nifty_df, cfg.get("fall", 10.0)) \
        if (cycle_df is not None and nifty_df is not None) else pd.DataFrame()
    if not paths.empty:
        pp = paths.copy()
        pp["date"] = pp["date"].map(_fmt_date)
        pp = pp.rename(columns=lambda c: c.replace("_", " ").title())
        _title(ws10, "NIFTY Cycle Path — day-by-day (running peak, drawdown, BUY/PEAK/FALL/SELL)",
               len(pp.columns))
        _write_df(ws10, pp, 3, green_col="Pct From Buy", max_rows=20000)
    else:
        _title(ws10, "NIFTY Cycle Path", 2)
        ws10.cell(row=3, column=1, value="(pass nifty data + cycles to populate)")

    # ══ ⑪ Yearwise (quartile-breakdown, mirrors short-term system) ═══════════════
    ws11 = wb.create_sheet("Yearwise Summary")
    if per_trade_df is not None and not per_trade_df.empty:
        _title(ws11, "Yearwise Output & Summary — bought stocks by entry year  ·  "
               "Quartiles from Eligible Stock Ranking (Q4 = best 25%)", 17)
        dfq = _attach_qnum(per_trade_df, eligible_ranks_df)
        nxt = _write_summary_q(ws11, 3, "Performance Summary by Period",
                               _summary_rows_q(dfq, "period"))
        _write_summary_q(ws11, nxt + 1, "Yearwise Breakdown",
                         _summary_rows_q(dfq, "year"))
        # plain equity-style yearwise (from investment analysis) appended below
        if ia is not None and not ia["yearwise"].empty:
            yw = ia["yearwise"].copy()
            yw["Year"] = yw["Year"].astype(str)
            for c in ["Invested", "Profit", "Cumulative Equity"]:
                if c in yw.columns:
                    yw[c] = yw[c].round(0)
            for c in ["Return%", "Avg Alpha", "Win Rate"]:
                if c in yw.columns:
                    yw[c] = yw[c].round(2)
            r2 = ws11.max_row + 2
            _band(ws11, r2, "Capital / Equity by Year", len(yw.columns))
            _write_df(ws11, yw, r2 + 1, green_col="Profit", bold_total_last=True)
    else:
        _title(ws11, "Yearwise Summary", 2)
        ws11.cell(row=3, column=1, value="(no trades)")

    # ══ ⑫ Phase Schedule ══════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("Phase Schedule")
    if phases is not None and not phases.empty:
        ph = phases.copy()
        for dc in ["entry_date", "exit_date"]:
            if dc in ph.columns:
                ph[dc] = ph[dc].map(_fmt_date)
        ph = ph.rename(columns=lambda c: c.replace("_", " ").title())
        _title(ws12, "Phase Schedule", len(ph.columns))
        _write_df(ws12, ph, 3)
    else:
        _title(ws12, "Phase Schedule", 2)
        ws12.cell(row=3, column=1, value="(no phases)")

    # ══ Per-cycle DETAIL sheets (linked from the Cycle Ledger) ═══════════════════
    if detail_cycles:
        for cval in detail_cycles:
            try:
                _build_cycle_detail(
                    wb, cval, cycle_df, per_trade_df, candidates_df,
                    eligible_ranks_df, leg_rank_df, paths, ia,
                    initial_capital, reinvest)
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
