"""
S³ Short (multi-leg) System — INTERACTIVE Excel export
======================================================
Mirrors the momentum interactive workbook: a Dashboard with a WINDOW selector
that drives live (INDEX/MATCH/COUNTIF) filtered views of the trades, the common
candidates and the per-window stock ranking for that window — plus static,
highlighted overview sheets (all trades, per-stock summary, window status).

Re-uses the generic table/filter/format helpers from `momentum_interactive`
so the two exports stay visually and behaviourally consistent. The only
difference is the filter key: the momentum book keys on "Cycle", this one keys
on "Window".
"""
from __future__ import annotations

import io
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

try:
    from export.momentum_interactive import (
        _df_to_table, _filter_view, _kv, _title, _band, _static_table,
        _fmt_date, _co, _FONT, _HEAD_BG, _HEAD_FG, _TITLE_BG, _CTRL,
    )
except Exception:  # pragma: no cover
    from momentum_interactive import (  # type: ignore
        _df_to_table, _filter_view, _kv, _title, _band, _static_table,
        _fmt_date, _co, _FONT, _HEAD_BG, _HEAD_FG, _TITLE_BG, _CTRL,
    )

try:
    from core.investment_analysis import compute_investment_analysis, fmt_inr
except Exception:  # pragma: no cover
    from investment_analysis import compute_investment_analysis, fmt_inr  # type: ignore


# ─────────────────────────────────────────────────────────────────────────────
# Data preparation — clean, tokenised display frames keyed on "Window"
# ─────────────────────────────────────────────────────────────────────────────

def _winkey(df: pd.DataFrame) -> pd.Series:
    """1-based window number for any short-system frame."""
    if df is None or df.empty:
        return pd.Series([], dtype=int)
    if "window" in df.columns:
        return pd.to_numeric(df["window"], errors="coerce").fillna(0).astype(int)
    if "window_idx" in df.columns:
        return (pd.to_numeric(df["window_idx"], errors="coerce").fillna(-1).astype(int) + 1)
    return pd.Series(range(1, len(df) + 1), index=df.index)


def _yn(s) -> str:
    return "Y" if bool(s) else ""


def _prep(per_trade_df, candidates_df, window_status_df, window_ranks_df, summary_df):
    out = {}

    # ── Trades ────────────────────────────────────────────────────────────────
    if per_trade_df is not None and not per_trade_df.empty:
        t = per_trade_df.copy()
        tr = pd.DataFrame({
            "Window":      _winkey(t),
            "Ticker":      t.get("ticker"),
            "EntryDate":   t.get("entry_date").map(_fmt_date) if "entry_date" in t else "",
            "ExitDate":    t.get("exit_date").map(_fmt_date) if "exit_date" in t else "",
            "EntryPrice":  t.get("entry_price"),
            "ExitPrice":   t.get("exit_price"),
            "ReturnPct":   pd.to_numeric(t.get("return_pct"), errors="coerce"),
            "NiftyReturn": pd.to_numeric(t.get("nifty_return"), errors="coerce"),
            "Alpha":       pd.to_numeric(t.get("alpha"), errors="coerce"),
            "DaysHeld":    t.get("days_held"),
        })
        out["trades"] = tr.sort_values(["Window", "Ticker"]).reset_index(drop=True)
    else:
        out["trades"] = pd.DataFrame(columns=["Window", "Ticker", "EntryDate", "ExitDate",
                                              "EntryPrice", "ExitPrice", "ReturnPct",
                                              "NiftyReturn", "Alpha", "DaysHeld"])

    # ── Candidates (common stocks per window) ──────────────────────────────────
    if candidates_df is not None and not candidates_df.empty:
        c = candidates_df.copy()
        cand = pd.DataFrame({
            "Window":     _winkey(c),
            "CommonRank": c.get("common_rank"),
            "Ticker":     c.get("ticker"),
            "Quartile":   c.get("quartile"),
            "MeanAlpha":  pd.to_numeric(c.get("mean_alpha"), errors="coerce"),
            "MeanReturn": pd.to_numeric(c.get("mean_return"), errors="coerce"),
            "Selected":   c.get("selected_for_trade").map(_yn) if "selected_for_trade" in c else "",
            "Traded":     c.get("traded").map(_yn) if "traded" in c else "",
        })
        out["cands"] = cand.sort_values(["Window", "CommonRank"]).reset_index(drop=True)
    else:
        out["cands"] = pd.DataFrame(columns=["Window", "CommonRank", "Ticker", "Quartile",
                                             "MeanAlpha", "MeanReturn", "Selected", "Traded"])

    # ── Per-window full stock ranking ──────────────────────────────────────────
    if window_ranks_df is not None and not window_ranks_df.empty:
        wr = window_ranks_df.copy()
        rank_col = next((x for x in ["rank", "window_rank", "common_rank"] if x in wr.columns), None)
        ret_col  = next((x for x in ["return_pct", "ret"] if x in wr.columns), None)
        rk = pd.DataFrame({
            "Window":    _winkey(wr),
            "Rank":      wr.get(rank_col) if rank_col else range(1, len(wr) + 1),
            "Ticker":    wr.get("ticker"),
            "Alpha":     pd.to_numeric(wr.get("alpha"), errors="coerce") if "alpha" in wr else np.nan,
            "ReturnPct": pd.to_numeric(wr.get(ret_col), errors="coerce") if ret_col else np.nan,
            "Common":    wr.get("common").map(_yn) if "common" in wr else (
                         wr.get("is_common").map(_yn) if "is_common" in wr else ""),
            "Traded":    wr.get("traded").map(_yn) if "traded" in wr else "",
        })
        out["ranks"] = rk.sort_values(["Window", "Rank"]).reset_index(drop=True)
    else:
        out["ranks"] = pd.DataFrame(columns=["Window", "Rank", "Ticker", "Alpha",
                                             "ReturnPct", "Common", "Traded"])

    # ── Window status (static overview) ────────────────────────────────────────
    if window_status_df is not None and not window_status_df.empty:
        s = window_status_df.copy()
        st_ = pd.DataFrame({
            "Window":    _winkey(s),
            "Status":    s.get("status"),
            "Trades":    s.get("n_trades"),
            "Reshuffled": s.get("reshuffled").map(_yn) if "reshuffled" in s else "",
            "FallDD%":   pd.to_numeric(s.get("fall_drawdown_pct"), errors="coerce") if "fall_drawdown_pct" in s else np.nan,
            "Reason":    s.get("reason"),
        })
        out["status"] = st_.sort_values("Window").reset_index(drop=True)
    else:
        out["status"] = pd.DataFrame(columns=["Window", "Status", "Trades", "Reshuffled", "FallDD%", "Reason"])

    # ── Per-stock summary (static overview) ────────────────────────────────────
    out["summary"] = summary_df.copy() if (summary_df is not None and not summary_df.empty) else pd.DataFrame()

    # window list + per-window max-row counts (for the filter views)
    wins = sorted(set(int(x) for x in out["trades"]["Window"].tolist()
                      + out["cands"]["Window"].tolist()
                      + out["status"]["Window"].tolist()
                      + out["ranks"]["Window"].tolist()))
    out["windows"] = wins
    return out


def _kmax(df: pd.DataFrame) -> int:
    if df is None or df.empty or "Window" not in df.columns:
        return 1
    vc = df["Window"].value_counts()
    return int(vc.max()) if len(vc) else 1


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def generate_short_interactive_excel(
    per_trade_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    window_status_df: pd.DataFrame,
    candidates_df: pd.DataFrame,
    window_ranks_df: pd.DataFrame = None,
    config: dict = None,
    initial_capital: float = 100_000.0,
    alloc_mode: str = "equal",
    reinvest: bool = False,
) -> bytes:
    cfg = config or {}
    data = _prep(per_trade_df, candidates_df, window_status_df, window_ranks_df, summary_df)
    wins = data["windows"] or [1]

    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"

    # hidden data tables (keyed on Window)
    _df_to_table(wb, "_TradesData", data["trades"], "tblTrades", key_col="Window")
    _df_to_table(wb, "_CandData",   data["cands"],  "tblCands",  key_col="Window")
    _df_to_table(wb, "_RankData",   data["ranks"],  "tblRanks",  key_col="Window")

    # ── Dashboard header + window selector ─────────────────────────────────────
    _title(dash, "S³ Short System — Interactive Window Explorer", 8)
    dash.cell(row=2, column=1, value="Select Window ▶").font = Font(name=_FONT, bold=True, size=11)
    ctrl = dash.cell(row=2, column=3, value=wins[0])
    ctrl.font = Font(name=_FONT, bold=True, size=12, color="1A1A1A")
    ctrl.fill = PatternFill("solid", fgColor=_CTRL)
    ctrl.alignment = Alignment(horizontal="center")
    dv = DataValidation(type="list", formula1=f'"{",".join(str(w) for w in wins)}"', allow_blank=False)
    dash.add_data_validation(dv); dv.add(ctrl)
    dash.cell(row=2, column=4, value="◀ pick a window; all tables below update")\
        .font = Font(name=_FONT, italic=True, size=9, color="888888")

    CTRL = "Dashboard!$C$2"

    def mi(col):
        return f'=IFERROR(INDEX(tblTrades[{col}],MATCH({CTRL},tblTrades[Window],0)),"—")'

    # ── This-window summary (live formulas) ────────────────────────────────────
    _band(dash, 4, "Selected Window — Summary", 8)
    pairs = [
        ("Stocks traded", f"=COUNTIFS(tblTrades[Window],{CTRL})"),
        ("Avg return %",  f'=IFERROR(AVERAGEIFS(tblTrades[ReturnPct],tblTrades[Window],{CTRL}),"—")'),
        ("Avg alpha %",   f'=IFERROR(AVERAGEIFS(tblTrades[Alpha],tblTrades[Window],{CTRL}),"—")'),
        ("Win rate %",    f'=IFERROR(COUNTIFS(tblTrades[Window],{CTRL},tblTrades[ReturnPct],">0")'
                          f'/COUNTIFS(tblTrades[Window],{CTRL})*100,"—")'),
        ("Best return %", f'=IFERROR(MAXIFS(tblTrades[ReturnPct],tblTrades[Window],{CTRL}),"—")'),
        ("Worst return %", f'=IFERROR(MINIFS(tblTrades[ReturnPct],tblTrades[Window],{CTRL}),"—")'),
        ("Avg days held", f'=IFERROR(AVERAGEIFS(tblTrades[DaysHeld],tblTrades[Window],{CTRL}),"—")'),
        ("Common candidates", f"=COUNTIFS(tblCands[Window],{CTRL})"),
    ]
    r = _kv(dash, 5, pairs)

    # ── Filtered views ─────────────────────────────────────────────────────────
    kmax_tr = _kmax(data["trades"]); kmax_cd = _kmax(data["cands"]); kmax_rk = _kmax(data["ranks"])

    r += 1
    _band(dash, r, "Trades in this window (green = profit, red = loss)", 10); r += 1
    _filter_view(dash, r,
                 ["Ticker", "Entry Date", "Exit Date", "Entry", "Exit", "Return %",
                  "NIFTY %", "Alpha", "Days"],
                 "tblTrades",
                 ["Ticker", "EntryDate", "ExitDate", "EntryPrice", "ExitPrice", "ReturnPct",
                  "NiftyReturn", "Alpha", "DaysHeld"],
                 kmax_tr, CTRL, green_col="Return %", key_col="Window")
    r += kmax_tr + 2

    _band(dash, r, "Common candidates (gold = traded)", 10); r += 1
    _filter_view(dash, r,
                 ["Common Rank", "Ticker", "Quartile", "Mean Alpha", "Mean Return", "Selected", "Traded"],
                 "tblCands",
                 ["CommonRank", "Ticker", "Quartile", "MeanAlpha", "MeanReturn", "Selected", "Traded"],
                 kmax_cd, CTRL, traded_col="Traded", key_col="Window")
    r += kmax_cd + 2

    if not data["ranks"].empty:
        _band(dash, r, "Per-window stock ranking (gold = traded)", 10); r += 1
        _filter_view(dash, r,
                     ["Rank", "Ticker", "Alpha", "Return %", "Common", "Traded"],
                     "tblRanks",
                     ["Rank", "Ticker", "Alpha", "ReturnPct", "Common", "Traded"],
                     kmax_rk, CTRL, green_col="Return %", traded_col="Traded", key_col="Window")
        r += kmax_rk + 2

    dash.column_dimensions["A"].width = 18
    dash.sheet_view.showGridLines = False

    # ── Config sheet ───────────────────────────────────────────────────────────
    wsC = wb.create_sheet("Run Settings")
    _title(wsC, "Run Configuration", 2)
    side = cfg.get("selection_side", "top")
    cfg_pairs = [
        ("Pattern", " → ".join(cfg.get("pattern", []))),
        ("Top N per leg", cfg.get("top_n")),
        ("Common stocks (K)", cfg.get("top_k_common")),
        ("Rank by", cfg.get("sort_by", "alpha")),
        ("Selection side", "Top-N (best) & common" if side == "top" else "Bottom-N (worst) & common"),
        ("Entry threshold %", cfg.get("entry_threshold_pct")),
        ("Exit threshold %", cfg.get("exit_threshold_pct")),
        ("Entry segment included", "Yes" if cfg.get("include_entry_segment", True) else "No"),
        ("Window volatility filter", cfg.get("winvol_mode", "off")),
        ("Initial capital (₹)", initial_capital),
        ("Reinvest profits", "Yes" if reinvest else "No"),
    ]
    _kv(wsC, 3, cfg_pairs)

    # ── Static: All Trades (highlighted) ───────────────────────────────────────
    wsT = wb.create_sheet("All Trades")
    _title(wsT, "All Trades — every executed trade (green = profit, red = loss)",
           len(data["trades"].columns))
    _static_table(wsT, data["trades"], 3, green_col="ReturnPct")

    # ── Static: Per-Stock Summary ──────────────────────────────────────────────
    if not data["summary"].empty:
        wsS = wb.create_sheet("Stock Summary")
        disp = data["summary"].rename(columns=lambda c: str(c).replace("_", " ").title())
        _title(wsS, "Per-Stock Summary", len(disp.columns))
        _static_table(wsS, disp, 3,
                      green_col=("Avg Return" if "Avg Return" in disp.columns else None))

    # ── Static: Window Status ──────────────────────────────────────────────────
    wsW = wb.create_sheet("Window Status")
    _title(wsW, "Per-Window Status", len(data["status"].columns))
    _static_table(wsW, data["status"], 3)

    # Dashboard is created first (active), so it already leads the workbook.
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
